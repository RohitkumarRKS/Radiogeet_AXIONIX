import json
import random
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone

from .forms import SignupForm, LoginForm, WorkspaceSetupForm
from .models import UserProfile, WorkspaceConfig, Tank, TankReading, Alarm, ModulePermission, EventLog, FlowMeter, FlowMeterReading, LicenseConfig


def splash_view(request):
    """Page 1: Splash/Loading screen."""
    return render(request, "splash.html")


from functools import wraps

LIFETIME_KEYS = {
    "RADIOGEET-AXIONIX-S3CR3T-K3Y-2026",
}

def ensure_one_year_keys_exist():
    try:
        from .models import LicenseKey
        one_year_keys = [
            "RADIOGEET-AXIONIX-KEY-2026-01",
            "RADIOGEET-AXIONIX-KEY-2026-02",
            "RADIOGEET-AXIONIX-KEY-2026-03",
            "RADIOGEET-AXIONIX-KEY-2026-04",
            "RADIOGEET-AXIONIX-KEY-2026-05",
            "RADIOGEET-AXIONIX-KEY-2026-06",
            "RADIOGEET-AXIONIX-KEY-2026-07",
            "RADIOGEET-AXIONIX-KEY-2026-08",
            "RADIOGEET-AXIONIX-KEY-2026-09",
            "RADIOGEET-AXIONIX-KEY-2026-10",
        ]
        for key in one_year_keys:
            LicenseKey.objects.get_or_create(key=key, defaults={"duration_days": 365})

        two_year_keys = [
            "RADIOGEET-SRKB-TDH6-SZ3J",
            "RADIOGEET-ZSQI-57SS-4D0Y",
            "RADIOGEET-ODL1-V46Y-BL48",
            "RADIOGEET-AA8P-782N-IYXS",
            "RADIOGEET-BMHO-7TJN-BVY0",
            "RADIOGEET-PCYL-KJI5-3TNB",
            "RADIOGEET-WU61-8WP1-DYWS",
            "RADIOGEET-D1B9-SESZ-MT3N",
            "RADIOGEET-W7CS-4AZX-CPYO",
            "RADIOGEET-PJPM-GN2Z-XJ2D",
        ]
        for key in two_year_keys:
            LicenseKey.objects.get_or_create(key=key, defaults={"duration_days": 730})

        three_year_keys = [
            "RADIOGEET-LVDW-I90R-Z0CB",
            "RADIOGEET-N25I-HFAI-38YZ",
            "RADIOGEET-89I1-RCEQ-DBI3",
            "RADIOGEET-M2QD-01NY-FCJF",
            "RADIOGEET-SP4Q-3S0B-LQSV",
            "RADIOGEET-PHL8-FDPU-L91Q",
            "RADIOGEET-968P-TBYQ-IFQN",
            "RADIOGEET-9Z6X-KT3J-RJQ2",
            "RADIOGEET-MMF1-7M6C-44PZ",
            "RADIOGEET-C60H-J9SW-PVYS",
        ]
        for key in three_year_keys:
            LicenseKey.objects.get_or_create(key=key, defaults={"duration_days": 1095})

        lifetime_keys = [
            "RADIOGEET-2VJ0-JR04-OM5D",
            "RADIOGEET-9IQM-13FF-5KZA",
            "RADIOGEET-Q99Y-WYDF-KNIR",
            "RADIOGEET-QGHV-99JZ-C1W5",
            "RADIOGEET-MDQK-WD9R-10FV",
            "RADIOGEET-P4Z1-Y1QJ-0RHP",
            "RADIOGEET-RZ1R-Q0UN-8357",
            "RADIOGEET-D872-4YDZ-6RIU",
            "RADIOGEET-WHK8-24WO-AA1H",
            "RADIOGEET-Z5LP-NYJS-6MGV",
        ]
        for key in lifetime_keys:
            LicenseKey.objects.get_or_create(key=key, defaults={"duration_days": 36500})
    except Exception:
        pass

def check_license_expiry(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        ensure_one_year_keys_exist()
        from .models import LicenseKey
        config = WorkspaceConfig.objects.first()
        if config:
            if config.activation_key == 'RADIOGEET-AXIONIX-S3CR3T-K3Y-2026-3T':
                if config.created_at + timedelta(days=3) < timezone.now():
                    messages.error(request, "Your 3-Day Trial has expired. Please enter a valid lifetime license key to continue.")
                    return redirect('renew_license')
            elif config.activation_key not in LIFETIME_KEYS:
                try:
                    lic = LicenseKey.objects.get(key=config.activation_key)
                    if lic.expires_at and lic.expires_at < timezone.now():
                        duration_str = "1-Year"
                        if lic.duration_days == 730:
                            duration_str = "2-Year"
                        elif lic.duration_days == 1095:
                            duration_str = "3-Year"
                        elif lic.duration_days >= 36500:
                            duration_str = "Lifetime"
                        messages.error(request, f"Your {duration_str} License has expired. Please enter a new valid license key to continue.")
                        return redirect('renew_license')
                except LicenseKey.DoesNotExist:
                    messages.error(request, "Invalid License Key detected. Please enter a valid license key.")
                    return redirect('renew_license')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def get_user_tanks(user):
    """
    Returns the queryset of active tanks that the user is allowed to see.
    If the user is an admin/superuser, they see all active tanks.
    If they are a normal user, but have no assigned tanks, they see all active tanks by default.
    Otherwise, they only see their explicitly assigned tanks.
    """
    is_admin = False
    try:
        is_admin = (user.is_superuser or user.profile.role == "admin")
    except:
        pass

    if is_admin:
        return Tank.objects.filter(is_active=True)
    
    try:
        profile = user.profile
        if profile.assigned_tanks.exists():
            return profile.assigned_tanks.filter(is_active=True)
    except:
        pass
        
    return Tank.objects.filter(is_active=True)

def renew_license_view(request):
    ensure_one_year_keys_exist()
    from .models import LicenseKey
    config = WorkspaceConfig.objects.first()
    
    if request.method == "POST":
        new_key = request.POST.get("activation_key", "").strip()
        
        if new_key in LIFETIME_KEYS or new_key == "RADIOGEET-AXIONIX-S3CR3T-K3Y-2026-3T":
            if config:
                config.activation_key = new_key
                config.save()
            messages.success(request, "License successfully updated! You can now login.")
            return redirect("login")
            
        try:
            lic = LicenseKey.objects.get(key=new_key)
            if not lic.is_used:
                lic.is_used = True
                lic.activated_at = timezone.now()
                lic.expires_at = lic.activated_at + timedelta(days=lic.duration_days)
                lic.save()
                
                if config:
                    config.activation_key = new_key
                    config.save()
                
                duration_str = "1 Year"
                if lic.duration_days == 730:
                    duration_str = "2 Years"
                elif lic.duration_days == 1095:
                    duration_str = "3 Years"
                elif lic.duration_days >= 36500:
                    duration_str = "Lifetime"
                messages.success(request, f"License successfully renewed for {duration_str}! You can now login.")
                return redirect("login")
            else:
                messages.error(request, "This license key has already been activated.")
        except LicenseKey.DoesNotExist:
            messages.error(request, "Invalid Activation Key. Please enter a valid, active Radiogeet license key.")
            
    return render(request, "auth/renew_license.html")


def welcome_view(request):
    """Page 2: Welcome/Workspace setup."""
    ensure_one_year_keys_exist()
    config = WorkspaceConfig.objects.first()
    
    # Check if the license has expired
    is_expired = False
    if config:
        if config.activation_key == 'RADIOGEET-AXIONIX-S3CR3T-K3Y-2026-3T':
            if config.created_at + timedelta(days=3) < timezone.now():
                is_expired = True
        elif config.activation_key not in LIFETIME_KEYS:
            from .models import LicenseKey
            try:
                lic = LicenseKey.objects.get(key=config.activation_key)
                if lic.expires_at and lic.expires_at < timezone.now():
                    is_expired = True
            except LicenseKey.DoesNotExist:
                is_expired = True

    # If the system is already configured and activated, redirect directly to login
    # unless the license has expired (in which case they must enter a valid key)
    if config and config.is_activated and not is_expired:
        return redirect("login")
    elif config and config.is_activated and is_expired:
        return redirect("renew_license")

    if request.method == "POST":
        was_activated = config and config.is_activated
        form = WorkspaceSetupForm(request.POST, instance=config)
        if form.is_valid():
            workspace = form.save(commit=False)
            workspace.is_activated = True  # Setup complete and activated
            workspace.save()
            
            # Activate the dynamic license key if applicable
            from .models import LicenseKey
            try:
                lic = LicenseKey.objects.get(key=workspace.activation_key)
                if not lic.is_used:
                    lic.is_used = True
                    lic.activated_at = timezone.now()
                    lic.expires_at = lic.activated_at + timedelta(days=lic.duration_days)
                    lic.save()
            except LicenseKey.DoesNotExist:
                pass
                
            messages.success(request, "Workspace setup initialized successfully!")
            if was_activated:
                return redirect("login")
            else:
                return redirect("signup")
    else:
        form = WorkspaceSetupForm(instance=config)
    return render(request, "welcome.html", {"form": form})


def signup_view(request):
    """Page 3: Create new account."""
    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            # Create user
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            # Split full name
            name_parts = form.cleaned_data["full_name"].split(" ", 1)
            user.first_name = name_parts[0]
            user.last_name = name_parts[1] if len(name_parts) > 1 else ""
            user.save()

            # Create profile
            UserProfile.objects.create(
                user=user,
                mobile_number=form.cleaned_data.get("mobile_number", ""),
                role=form.cleaned_data["role"],
                organization=form.cleaned_data.get("organization", ""),
                timezone=form.cleaned_data.get("timezone", "Asia/Kolkata"),
            )

            messages.success(request, "Account created successfully! Please login.")
            return redirect("login")
    else:
        form = SignupForm()
    return render(request, "auth/signup.html", {"form": form})


def login_view(request):
    """Page 4: Login page."""
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data["username_or_email"]
            password = form.cleaned_data["password"]
            remember_me = form.cleaned_data.get("remember_me", False)

            # Try to authenticate with username first, then email
            user = authenticate(request, username=username_or_email, password=password)
            if user is None:
                # Try email lookup
                try:
                    user_obj = User.objects.get(email=username_or_email)
                    user = authenticate(request, username=user_obj.username, password=password)
                except User.DoesNotExist:
                    user = None

            if user is not None:
                # Check if user has an active profile assigned by admin
                try:
                    profile = user.profile
                    if not profile.is_active_user:
                        messages.error(request, "Your account is currently disabled. Please contact the administrator.")
                        return redirect("login")
                except UserProfile.DoesNotExist:
                    if not user.is_superuser:
                        messages.error(request, "User profile not found. Please contact the administrator.")
                        return redirect("login")

                login(request, user)
                if not remember_me:
                    request.session.set_expiry(0)  # Session expires on browser close
                messages.success(request, f"Welcome back, {user.get_full_name() or user.username}!")
                return redirect("dashboard")
            else:
                messages.error(request, "Invalid username/email or password.")
    else:
        form = LoginForm()
    return render(request, "auth/login.html", {"form": form})


def logout_view(request):
    """Logout and redirect to login."""
    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect("login")


def format_tank_level(level_percent, capacity_kl, unit, raw_val=None):
    if level_percent is None and raw_val is None:
        return "--"
    unit = (unit or "%").strip()
    cap = float(capacity_kl) if capacity_kl else 5.0
    if unit == "RAW":
        v = raw_val if raw_val is not None else level_percent
        if v is None:
            return "--"
        if isinstance(v, (int, float)) and float(v).is_integer():
            return f"{int(v)}"
        return f"{float(v):.1f}"
    if level_percent is None:
        return "--"
    if unit == "%":
        return f"{level_percent:.1f}%"
    elif unit in ("L", "Liters"):
        liters = (level_percent / 100.0) * cap * 1000.0
        return f"{round(liters)} L"
    elif unit in ("mL", "Milliliters"):
        ml = (level_percent / 100.0) * (cap * 1000.0 if cap <= 100 else cap) * 1000.0
        return f"{round(ml)} mL"
    elif unit in ("KL", "Kiloliters"):
        kl = (level_percent / 100.0) * cap
        return f"{kl:.1f} KL"
    elif unit in ("m³", "m3"):
        m3 = (level_percent / 100.0) * cap
        return f"{m3:.1f} m³"
    elif unit in ("gal", "Gallons"):
        gal = (level_percent / 100.0) * cap * 264.172
        return f"{gal:.1f} gal"
    else:
        return f"{level_percent:.1f} {unit}"


def get_tank_value_in_unit(level_percent, capacity_kl, unit, raw_val=None):
    if unit == "RAW":
        return raw_val if raw_val is not None else (level_percent if level_percent is not None else 0.0)
    if level_percent is None:
        return 0.0
    cap = float(capacity_kl) if capacity_kl else 5.0
    if unit == "%":
        return level_percent
    elif unit in ("L", "Liters"):
        return (level_percent / 100.0) * cap * 1000.0
    elif unit in ("mL", "Milliliters"):
        return (level_percent / 100.0) * (cap * 1000.0 if cap <= 100 else cap) * 1000.0
    elif unit in ("KL", "Kiloliters"):
        return (level_percent / 100.0) * cap
    elif unit in ("m³", "m3"):
        return (level_percent / 100.0) * cap
    elif unit in ("gal", "Gallons"):
        return (level_percent / 100.0) * cap * 264.172
    else:
        return level_percent


def format_tank_capacity(capacity_kl, unit, capacity_unit=None):
    if capacity_kl is None:
        return "--"
    c_unit = (capacity_unit or unit or "%").strip()
    cap = float(capacity_kl)
    if c_unit in ("RAW", "None", "", "No Unit"):
        return f"{int(cap)}" if cap.is_integer() else f"{cap:.1f}"
    elif c_unit in ("L", "Liters"):
        return f"{round(cap)} L" if cap > 100 else f"{round(cap * 1000.0)} L"
    elif c_unit in ("mL", "Milliliters"):
        return f"{round(cap)} mL" if cap > 100 else f"{round(cap * 1000000.0)} mL"
    elif c_unit in ("KL", "Kiloliters", "%"):
        return f"{cap:.1f} KL"
    elif c_unit in ("m³", "m3"):
        return f"{cap:.1f} m³"
    elif c_unit in ("gal", "Gallons"):
        return f"{cap * 264.172:.1f} gal"
    else:
        return f"{cap:.1f} {c_unit}"


def format_total_val(val, unit):
    if val is None:
        return "--"
    try:
        v = float(val)
    except (ValueError, TypeError):
        return str(val)
    unit_str = (unit or "").strip()
    if unit_str in ("RAW", "None", "", "No Unit"):
        if v.is_integer():
            return f"{int(v)}"
        return f"{v:.1f}"
    elif unit_str in ("%", "Percent"):
        return f"{v:.1f}%"
    elif unit_str in ("L", "Liters"):
        return f"{round(v)} L" if abs(v) > 100 else f"{v:.1f} L"
    elif unit_str in ("mL", "Milliliters"):
        return f"{round(v)} mL"
    elif unit_str in ("KL", "Kiloliters"):
        return f"{v:.1f} KL"
    elif unit_str in ("m³", "m3"):
        return f"{v:.1f} m³"
    elif unit_str in ("gal", "Gallons"):
        return f"{v:.1f} gal"
    else:
        if v.is_integer():
            return f"{int(v)} {unit_str}"
        return f"{v:.1f} {unit_str}"



def get_data_changes(tanks, flow_meters, start_dt, end_dt):
    from .models import TankReading, FlowMeterReading
    changes = []
    
    # 1. Check changes for Tanks
    for tank in tanks:
        prev_reading = TankReading.objects.filter(tank=tank, timestamp__lt=start_dt).order_by('-timestamp').first()
        readings = TankReading.objects.filter(tank=tank, timestamp__range=(start_dt, end_dt)).order_by('timestamp')
        
        last_val = None
        if prev_reading:
            if tank.widget_type == 'flow_meter':
                last_val = (prev_reading.flow_rate, prev_reading.total_flow)
            else:
                last_val = (prev_reading.level_percent, prev_reading.raw_value)
                
        for r in readings:
            if tank.widget_type == 'flow_meter':
                curr_val = (r.flow_rate, r.total_flow)
                if last_val is None:
                    last_val = curr_val
                    continue
                if curr_val[0] != last_val[0] or curr_val[1] != last_val[1]:
                    old_str = f"Rate: {last_val[0]} {tank.flow_unit}, Vol: {last_val[1]} {tank.total_unit}"
                    new_str = f"Rate: {r.flow_rate} {tank.flow_unit}, Vol: {r.total_flow} {tank.total_unit}"
                    changes.append({
                        'timestamp': r.timestamp,
                        'device_id': tank.tank_id,
                        'device_name': tank.name,
                        'change_desc': f"Changed from [{old_str}] to [{new_str}]"
                    })
                    last_val = curr_val
            else:
                curr_val = (r.level_percent, r.raw_value)
                if last_val is None:
                    last_val = curr_val
                    continue
                if curr_val[0] != last_val[0]:
                    old_str = format_tank_level(last_val[0], tank.capacity_kl, tank.unit, last_val[1])
                    new_str = format_tank_level(r.level_percent, tank.capacity_kl, tank.unit, r.raw_value)
                    changes.append({
                        'timestamp': r.timestamp,
                        'device_id': tank.tank_id,
                        'device_name': tank.name,
                        'change_desc': f"Changed from [{old_str}] to [{new_str}]"
                    })
                    last_val = curr_val
                    
    # 2. Check changes for FlowMeters
    for fm in flow_meters:
        prev_reading = FlowMeterReading.objects.filter(flow_meter=fm, timestamp__lt=start_dt).order_by('-timestamp').first()
        readings = FlowMeterReading.objects.filter(flow_meter=fm, timestamp__range=(start_dt, end_dt)).order_by('timestamp')
        
        last_val = None
        if prev_reading:
            last_val = (prev_reading.flow_rate, prev_reading.total_volume)
            
        for r in readings:
            curr_val = (r.flow_rate, r.total_volume)
            if last_val is None:
                last_val = curr_val
                continue
            if curr_val[0] != last_val[0] or curr_val[1] != last_val[1]:
                old_str = f"Rate: {last_val[0]} {fm.flow_unit}, Vol: {last_val[1]} {fm.total_unit}"
                new_str = f"Rate: {r.flow_rate} {fm.flow_unit}, Vol: {r.total_volume} {fm.total_unit}"
                changes.append({
                    'timestamp': r.timestamp,
                    'device_id': fm.meter_id,
                    'device_name': fm.name,
                    'change_desc': f"Changed from [{old_str}] to [{new_str}]"
                })
                last_val = curr_val
                
    changes = sorted(changes, key=lambda x: x['timestamp'], reverse=True)
    return changes


def get_modbus_connection_status(profile_name="Tanks"):
    from .models import SerialConnectionConfig, TankReading, FlowMeterReading
    config = SerialConnectionConfig.objects.filter(profile_name=profile_name).first()
    if config and config.com_port == "SIMULATOR":
        return True, "Simulator Active"

    # Check for active recent telemetry readings in the database (within last 120 seconds)
    recent_time = timezone.now() - timedelta(seconds=120)
    if profile_name == "Tanks":
        if TankReading.objects.filter(timestamp__gte=recent_time).exists():
            return True, "Telemetry scan active"
    elif profile_name == "Flow Meters":
        if FlowMeterReading.objects.filter(timestamp__gte=recent_time).exists() or TankReading.objects.filter(tank__widget_type="flow_meter", timestamp__gte=recent_time).exists():
            return True, "Telemetry scan active"

    last_log = EventLog.objects.filter(source=f"Modbus Daemon ({profile_name})").order_by("-timestamp").first()
    if not last_log:
        last_log = EventLog.objects.filter(source__icontains="Modbus Daemon").order_by("-timestamp").first()

    is_connected = False
    connection_msg = "Disconnected"
    if last_log:
        connection_msg = last_log.message
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message or "active" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True
        elif (timezone.now() - last_log.timestamp).total_seconds() <= 120 and "unavailable" not in last_log.message:
            is_connected = True

    return is_connected, connection_msg


@login_required
@check_license_expiry
def dashboard_view(request):
    """Page 5: Tank Level Monitoring Dashboard."""
    # Get last Modbus daemon status log for Tanks profile to verify active connection
    is_connected, connection_msg = get_modbus_connection_status("Tanks")

    # Get available COM ports
    import serial.tools.list_ports
    try:
        available_ports = [p.device for p in serial.tools.list_ports.comports()]
    except Exception:
        available_ports = []

    from .models import SerialConnectionConfig
    config_obj = SerialConnectionConfig.objects.filter(profile_name="Tanks").first()
    tanks_com = config_obj.com_port if config_obj else "SIMULATOR"
    tanks_baud = config_obj.baud_rate if config_obj else 9600
    tanks_parity = config_obj.parity if config_obj else "None"

    # Determine which tanks to show based on user profile assignment
    tanks = get_user_tanks(request.user).exclude(widget_type="flow_meter").order_by("tank_id")

    tank_data = []
    high_alarms_tanks = 0
    low_alarms_tanks = 0
    total_level = 0
    valid_tanks = 0

    for tank in tanks:
        latest = tank.readings.first()
        if latest:
            level = latest.level_percent
            clamped_level = max(0.0, min(100.0, level))
            level_dasharray = round(clamped_level * 0.75, 1)
            # Determine status based on tank thresholds
            if level >= tank.high_limit:
                status = "High Alarm"
                status_class = "alarm-high"
                high_alarms_tanks += 1
            elif level <= tank.low_limit:
                status = "Low Alarm"
                status_class = "alarm-low"
                low_alarms_tanks += 1
            else:
                status = "Normal"
                status_class = "normal"
            total_level += level
            valid_tanks += 1
        else:
            level = None
            clamped_level = 0.0
            level_dasharray = 0
            status = "No Data"
            status_class = "no-data"

        raw_val = getattr(latest, 'raw_value', None) if latest else None
        formatted_lvl = format_tank_level(level, tank.capacity_kl, tank.unit, raw_val=raw_val)
        formatted_cap = format_tank_capacity(tank.capacity_kl, tank.unit, getattr(tank, 'capacity_unit', 'KL'))

        tank_data.append({
            "id": tank.tank_id,
            "name": tank.name,
            "level": level,
            "raw_value": raw_val,
            "clamped_level": clamped_level,
            "formatted_level": formatted_lvl,
            "formatted_capacity": formatted_cap,
            "level_dasharray": level_dasharray,
            "status": status,
            "status_class": status_class,
            "capacity": tank.capacity_kl,
            "capacity_unit": getattr(tank, 'capacity_unit', 'KL') or 'KL',
            "widget_type": tank.widget_type,
            "register_address": tank.register_address,
            "high_limit": tank.high_limit,
            "low_limit": tank.low_limit,
            "slave_id": tank.slave_id,
            "function_code": tank.function_code,
            "unit": tank.unit,
            "raw_zero": tank.raw_zero,
            "raw_span": tank.raw_span,
            "scanner_raw_zero": tank.scanner_raw_zero,
            "scanner_raw_span": tank.scanner_raw_span,
            "com_port": tanks_com,
            "baud_rate": tanks_baud,
            "parity": tanks_parity,
            "error_accuracy": tank.error_accuracy,
        })

    total_tanks = len(tank_data)
    avg_level = round(total_level / valid_tanks, 1) if valid_tanks > 0 else None
    total_capacity_kl = sum(t["capacity"] for t in tank_data)
    l_tanks = [t for t in tank_data if t["unit"] in ("L", "Liters")]
    gal_tanks = [t for t in tank_data if t["unit"] in ("gal", "Gallons")]
    m3_tanks = [t for t in tank_data if t["unit"] in ("m³", "m3")]

    if len(l_tanks) == len(tank_data) and len(tank_data) > 0:
        liters = round(total_capacity_kl * 1000.0)
        if liters >= 1_000_000:
            formatted_total_capacity = f"{liters / 1_000_000:.2f}M L"
        else:
            formatted_total_capacity = f"{liters:,} L"
    elif len(gal_tanks) == len(tank_data) and len(tank_data) > 0:
        gals = total_capacity_kl * 264.172
        if gals >= 1_000_000:
            formatted_total_capacity = f"{gals / 1_000_000:.2f}M gal"
        else:
            formatted_total_capacity = f"{gals:,.1f} gal"
    elif len(m3_tanks) == len(tank_data) and len(tank_data) > 0:
        if total_capacity_kl >= 1_000_000:
            formatted_total_capacity = f"{total_capacity_kl / 1_000_000:.2f}M m³"
        else:
            formatted_total_capacity = f"{total_capacity_kl:,.1f} m³"
    else:
        if total_capacity_kl >= 1_000_000:
            formatted_total_capacity = f"{total_capacity_kl / 1_000_000:.2f}M KL"
        elif total_capacity_kl >= 100_000:
            formatted_total_capacity = f"{total_capacity_kl / 1_000:.1f}k KL"
        elif total_capacity_kl >= 10_000:
            formatted_total_capacity = f"{total_capacity_kl:,.0f} KL"
        else:
            liters = round(total_capacity_kl * 1000.0)
            if liters >= 100_000:
                formatted_total_capacity = f"{total_capacity_kl:,.1f} KL"
            else:
                formatted_total_capacity = f"{int(total_capacity_kl)} KL ({liters:,} L)"

    # Get recent alarms
    recent_alarms = Alarm.objects.select_related("tank", "flow_meter").order_by("-timestamp")[:6]
    alarm_list = []
    for alarm in recent_alarms:
        tank_id = alarm.tank.tank_id if alarm.tank else (alarm.flow_meter.meter_id if alarm.flow_meter else "Unknown")
        tank_name = alarm.tank.name if alarm.tank else (alarm.flow_meter.name if alarm.flow_meter else "Unknown")
        alarm_list.append({
            "tank_id": tank_id,
            "tank_name": tank_name,
            "alarm_type": alarm.get_alarm_type_display(),
            "alarm_class": "alarm-high" if alarm.alarm_type == "high" else "alarm-low",
            "level": alarm.level_percent,
            "time": timezone.localtime(alarm.timestamp).strftime("%I:%M:%S %p"),
        })

    # Flow meter alarms breakdown
    from .models import FlowMeter
    high_alarms_fms = 0
    low_alarms_fms = 0
    active_fms = FlowMeter.objects.filter(is_active=True)
    for fm in active_fms:
        latest_fm = fm.readings.first()
        if latest_fm:
            if fm.high_limit and latest_fm.flow_rate >= fm.high_limit:
                high_alarms_fms += 1
            elif fm.low_limit and latest_fm.flow_rate <= fm.low_limit:
                low_alarms_fms += 1

    virtual_fms = Tank.objects.filter(is_active=True, widget_type="flow_meter")
    combined_flow_meters = []
    for fm in FlowMeter.objects.filter(is_active=True).order_by("meter_id"):
        combined_flow_meters.append({"meter_id": fm.meter_id})
    for fm in virtual_fms.order_by("tank_id"):
        combined_flow_meters.append({"meter_id": fm.tank_id})

    total_flow_meters = FlowMeter.objects.filter(is_active=True).count() + virtual_fms.count()
    total_devices = total_tanks + total_flow_meters
    total_high_alarms = high_alarms_tanks + high_alarms_fms
    total_low_alarms = low_alarms_tanks + low_alarms_fms

    context = {
        "tanks": tank_data,
        "flow_meters": combined_flow_meters,
        "total_tanks": total_tanks,
        "total_flow_meters": total_flow_meters,
        "total_devices": total_devices,
        "avg_level": avg_level,
        "total_capacity": formatted_total_capacity,
        "high_alarms": total_high_alarms,
        "low_alarms": total_low_alarms,
        "high_alarms_tanks": high_alarms_tanks,
        "high_alarms_fms": high_alarms_fms,
        "low_alarms_tanks": low_alarms_tanks,
        "low_alarms_fms": low_alarms_fms,
        "recent_alarms": alarm_list,
        "current_time": timezone.now(),
        "user": request.user,
        "is_connected": is_connected,
        "connection_msg": connection_msg,
        "available_ports": available_ports,
    }
    return render(request, "core/dashboard.html", context)


@login_required
def api_historical_data(request):
    """API endpoint returning real historical database readings for trend charts."""
    # Determine active tanks
    tanks = get_user_tanks(request.user).order_by("tank_id")

    # Get active flow meters
    flow_meters = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    # Filter by specific tank/meter if requested
    tank_id = request.GET.get("tank_id")
    if tank_id:
        if tank_id.startswith("FM-") or FlowMeter.objects.filter(meter_id=tank_id).exists():
            flow_meters = flow_meters.filter(meter_id=tank_id)
            tanks = Tank.objects.none()
        else:
            tanks = tanks.filter(tank_id=tank_id)
            flow_meters = FlowMeter.objects.none()

    # Filter by range parameter (1H, 6H, 12H, 24H, 7D, CUSTOM)
    range_val = request.GET.get("range", "6H").upper()
    now = timezone.now()
    
    start_time = None
    end_time = None
    
    if range_val == "CUSTOM":
        start_str = request.GET.get("start_time")
        end_str = request.GET.get("end_time")
        
        if start_str:
            try:
                naive_start = datetime.strptime(start_str, "%Y-%m-%dT%H:%M")
                start_time = timezone.make_aware(naive_start, timezone.get_current_timezone())
            except ValueError:
                pass
        if end_str:
            try:
                naive_end = datetime.strptime(end_str, "%Y-%m-%dT%H:%M")
                end_time = timezone.make_aware(naive_end, timezone.get_current_timezone())
            except ValueError:
                pass
                
        # Default fallback for custom if parsing fails
        if not start_time:
            start_time = now - timedelta(hours=6)
        if not end_time:
            end_time = now
    else:
        end_time = now
        if range_val == "1H":
            start_time = now - timedelta(hours=1)
        elif range_val == "12H":
            start_time = now - timedelta(hours=12)
        elif range_val == "24H":
            start_time = now - timedelta(hours=24)
        elif range_val == "7D":
            start_time = now - timedelta(days=7)
        else:
            start_time = now - timedelta(hours=6)

    # Query tanks with strict boundaries
    db_readings_qs = TankReading.objects.filter(tank__in=tanks).select_related("tank")
    if start_time:
        db_readings_qs = db_readings_qs.filter(timestamp__gte=start_time)
    if end_time:
        db_readings_qs = db_readings_qs.filter(timestamp__lte=end_time)
        
    db_readings_qs = db_readings_qs.order_by("-timestamp")
    
    total_db_count = TankReading.objects.filter(tank__in=tanks).count()
    if total_db_count == 0:
        db_readings = list(TankReading.objects.filter(tank__in=tanks).select_related("tank").order_by("-timestamp")[:150])
    else:
        db_readings = list(db_readings_qs[:300])
    
    db_readings.reverse()

    # Query flow meters with strict boundaries
    fm_readings_qs = FlowMeterReading.objects.filter(flow_meter__in=flow_meters).select_related("flow_meter")
    if start_time:
        fm_readings_qs = fm_readings_qs.filter(timestamp__gte=start_time)
    if end_time:
        fm_readings_qs = fm_readings_qs.filter(timestamp__lte=end_time)
        
    fm_readings_qs = fm_readings_qs.order_by("-timestamp")
    
    total_fm_count = FlowMeterReading.objects.filter(flow_meter__in=flow_meters).count()
    if total_fm_count == 0:
        fm_readings = list(FlowMeterReading.objects.filter(flow_meter__in=flow_meters).select_related("flow_meter").order_by("-timestamp")[:150])
    else:
        fm_readings = list(fm_readings_qs[:300])
    
    fm_readings.reverse()

    # Compile readings into timestamps and datasets with rounded 5-second grouping to align time slots
    time_groups = {}
    
    for r in db_readings:
        rounded_dt = r.timestamp
        seconds = round(rounded_dt.second / 5) * 5
        if seconds == 60:
            rounded_dt = rounded_dt + timedelta(minutes=1)
            seconds = 0
        rounded_dt = rounded_dt.replace(second=seconds, microsecond=0)
        local_dt = timezone.localtime(rounded_dt)
        if local_dt not in time_groups:
            time_groups[local_dt] = {}
        if r.tank.widget_type == "flow_meter":
            time_groups[local_dt][r.tank.tank_id] = r.flow_rate if r.flow_rate is not None else 0.0
        else:
            time_groups[local_dt][r.tank.tank_id] = get_tank_value_in_unit(r.level_percent, r.tank.capacity_kl, r.tank.unit, getattr(r, 'raw_value', None))

    for r in fm_readings:
        rounded_dt = r.timestamp
        seconds = round(rounded_dt.second / 5) * 5
        if seconds == 60:
            rounded_dt = rounded_dt + timedelta(minutes=1)
            seconds = 0
        rounded_dt = rounded_dt.replace(second=seconds, microsecond=0)
        local_dt = timezone.localtime(rounded_dt)
        if local_dt not in time_groups:
            time_groups[local_dt] = {}
        time_groups[local_dt][r.flow_meter.meter_id] = r.flow_rate

    # Sort the datetime keys chronologically
    sorted_dts = sorted(time_groups.keys())
    timestamps = [dt.isoformat() for dt in sorted_dts]
    
    # Format datasets for Chart.js
    datasets = {}
    for tank in tanks:
        datasets[tank.tank_id] = []
        for dt in sorted_dts:
            val = time_groups[dt].get(tank.tank_id, 0.0)
            datasets[tank.tank_id].append(val)

    for fm in flow_meters:
        datasets[fm.meter_id] = []
        for dt in sorted_dts:
            val = time_groups[dt].get(fm.meter_id, 0.0)
            datasets[fm.meter_id].append(val)

    return JsonResponse({
        "timestamps": timestamps,
        "datasets": datasets
    })


@login_required
def api_tank_data(request):
    """API endpoint for real-time tank data updates.
    Follows RadioDAQ pattern: if device is disconnected, return null/no-data for all tanks.
    Only show live data when the background daemon is actively reading registers successfully.
    """
    # Get last Modbus daemon status log for Tanks to verify active connection
    is_connected, _ = get_modbus_connection_status("Tanks")

    # Determine which tanks to show based on user profile assignment
    tanks = get_user_tanks(request.user).exclude(widget_type="flow_meter").order_by("tank_id")

    data = []
    for tank in tanks:
        # If device is disconnected, show null/no-data — do NOT show stale readings
        if not is_connected:
            data.append({
                "id": tank.tank_id,
                "name": tank.name,
                "level": None,
                "status": "Offline",
                "status_class": "no-data",
                "widget_type": tank.widget_type,
                "capacity": tank.capacity_kl,
                "unit": tank.unit,
                "formatted_level": "--",
                "formatted_capacity": format_tank_capacity(tank.capacity_kl, tank.unit),
            })
            continue

        # Device is connected — show live reading
        latest = tank.readings.first()
        raw_val = getattr(latest, 'raw_value', None) if latest else None
        if latest:
            level = latest.level_percent if latest.level_percent is not None else 0.0
            if level < 0 or level > 30000 or level in (32765, 32767, 65534, 65535):
                level = 0.0

            if raw_val is not None and (raw_val < 0 or raw_val > 30000 or raw_val in (32765, 32767, 65534, 65535)):
                raw_val = 0.0

            # Alarm check: if unit == 'RAW', limits are RAW values, compare against raw_val or level
            check_val = raw_val if (tank.unit == "RAW" and raw_val is not None) else level
            if check_val >= tank.high_limit:
                status = "High Alarm"
                status_class = "alarm-high"
            elif check_val <= tank.low_limit:
                status = "Low Alarm"
                status_class = "alarm-low"
            else:
                status = "Normal"
                status_class = "normal"
        else:
            level = None
            raw_val = None
            status = "No Data"
            status_class = "no-data"

        formatted_lvl = format_tank_level(level, tank.capacity_kl, tank.unit, raw_val=raw_val)
        formatted_cap = format_tank_capacity(tank.capacity_kl, tank.unit, getattr(tank, 'capacity_unit', 'KL'))

        data.append({
            "id": tank.tank_id,
            "name": tank.name,
            "level": level,
            "raw_value": raw_val,
            "status": status,
            "status_class": status_class,
            "widget_type": tank.widget_type,
            "capacity": tank.capacity_kl,
            "capacity_unit": getattr(tank, 'capacity_unit', 'KL') or 'KL',
            "unit": tank.unit,
            "formatted_level": formatted_lvl,
            "formatted_capacity": formatted_cap,
            "high_limit": tank.high_limit,
            "low_limit": tank.low_limit,
            "raw_zero": tank.raw_zero,
            "raw_span": tank.raw_span,
            "scanner_raw_zero": tank.scanner_raw_zero,
            "scanner_raw_span": tank.scanner_raw_span,
        })
    return JsonResponse({"tanks": data, "is_connected": is_connected, "timestamp": timezone.now().isoformat()})


@login_required
def api_flow_meter_data(request):
    """API endpoint for real-time flow meter data updates.
    Returns live flow rate and total volume readings from the Flow Meters daemon.
    Shows connection status independently from Tank daemon.
    """
    from .models import FlowMeter, FlowMeterReading, Tank

    # Check Flow Meter daemon status independently
    is_flow_connected, _ = get_modbus_connection_status("Flow Meters")

    # Check Tank daemon status (for virtual flow meters)
    is_tanks_connected, _ = get_modbus_connection_status("Tanks")

    # Determine which virtual flow meters to show based on user profile assignment
    virtual_fms = get_user_tanks(request.user).filter(widget_type="flow_meter").order_by("tank_id")

    flow_meters = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    from .models import SerialConnectionConfig
    tanks_config = SerialConnectionConfig.objects.filter(profile_name="Tanks").first()
    tanks_com = tanks_config.com_port if tanks_config else "SIMULATOR"
    flow_config = SerialConnectionConfig.objects.filter(profile_name="Flow Meters").first()
    flow_com = flow_config.com_port if flow_config else "SIMULATOR"

    data = []
    
    # 1. Add standard flow meters
    for fm in flow_meters:
        unit_flow = fm.flow_unit or "m³/h"
        unit_total = fm.total_unit or fm.flow_unit or "m³"

        if not is_flow_connected:
            data.append({
                "id": fm.meter_id,
                "name": fm.name,
                "flow_rate": None,
                "total_volume": None,
                "flow_unit": unit_flow,
                "total_unit": unit_total,
                "status": "Offline",
                "status_class": "no-data",
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": flow_com,
                "error_accuracy": fm.error_accuracy,
                "scanner_raw_zero": fm.scanner_raw_zero,
                "scanner_raw_span": fm.scanner_raw_span,
                "calibrated_span": fm.calibrated_span,
            })
            continue

        latest = fm.readings.first()
        if latest:
            flow_rate = latest.flow_rate if latest.flow_rate is not None else 0.0
            if flow_rate < 0 or flow_rate > 30000 or flow_rate in (32765, 32767, 65534, 65535):
                flow_rate = 0.0

            if flow_rate >= fm.high_limit:
                status = "High Alarm"
                status_class = "alarm-high"
            elif flow_rate <= fm.low_limit:
                status = "Low Alarm"
                status_class = "alarm-low"
            else:
                status = "Online"
                status_class = "normal"

            data.append({
                "id": fm.meter_id,
                "name": fm.name,
                "flow_rate": round(flow_rate, 3),
                "total_volume": round(latest.total_volume, 3),
                "flow_unit": unit_flow,
                "total_unit": unit_total,
                "status": status,
                "status_class": status_class,
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": flow_com,
                "error_accuracy": fm.error_accuracy,
                "scanner_raw_zero": fm.scanner_raw_zero,
                "scanner_raw_span": fm.scanner_raw_span,
                "calibrated_span": fm.calibrated_span,
            })
        else:
            data.append({
                "id": fm.meter_id,
                "name": fm.name,
                "flow_rate": None,
                "total_volume": None,
                "flow_unit": unit_flow,
                "total_unit": unit_total,
                "status": "No Data",
                "status_class": "no-data",
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": flow_com,
                "error_accuracy": fm.error_accuracy,
                "scanner_raw_zero": fm.scanner_raw_zero,
                "scanner_raw_span": fm.scanner_raw_span,
                "calibrated_span": fm.calibrated_span,
            })

    # 2. Add virtual flow meters
    for fm in virtual_fms:
        if not is_tanks_connected:
            data.append({
                "id": fm.tank_id,
                "name": fm.name,
                "flow_rate": None,
                "total_volume": None,
                "flow_unit": fm.flow_unit,
                "total_unit": fm.total_unit,
                "status": "Offline",
                "status_class": "no-data",
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": tanks_com,
                "error_accuracy": fm.error_accuracy,
            })
            continue

        latest = fm.readings.first()
        if latest and latest.flow_rate is not None:
            flow_rate = latest.flow_rate
            total_volume = latest.total_flow if latest.total_flow is not None else 0.0
            if flow_rate >= fm.high_limit:
                status = "High Alarm"
                status_class = "alarm-high"
            elif flow_rate <= fm.low_limit:
                status = "Low Alarm"
                status_class = "alarm-low"
            else:
                status = "Online"
                status_class = "normal"

            data.append({
                "id": fm.tank_id,
                "name": fm.name,
                "flow_rate": round(flow_rate, 3),
                "total_volume": round(total_volume, 3),
                "flow_unit": fm.flow_unit,
                "total_unit": fm.total_unit,
                "status": status,
                "status_class": status_class,
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": tanks_com,
                "error_accuracy": fm.error_accuracy,
            })
        else:
            data.append({
                "id": fm.tank_id,
                "name": fm.name,
                "flow_rate": None,
                "total_volume": None,
                "flow_unit": fm.flow_unit,
                "total_unit": fm.total_unit,
                "status": "No Data",
                "status_class": "no-data",
                "high_limit": fm.high_limit,
                "low_limit": fm.low_limit,
                "flow_rate_register": fm.flow_rate_register,
                "total_volume_register": fm.total_volume_register,
                "slave_id": fm.slave_id,
                "com_port": tanks_com,
                "error_accuracy": fm.error_accuracy,
            })

    return JsonResponse({
        "flow_meters": data,
        "is_connected": is_flow_connected or is_tanks_connected,
        "timestamp": timezone.now().isoformat()
    })


@login_required
@check_license_expiry
def trend_chart_view(request):
    """Page: Trend Chart."""
    # Get last Modbus daemon status log to verify active connection
    last_log = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    is_connected = False
    connection_msg = "Disconnected"
    if last_log:
        connection_msg = last_log.message
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True

    # Determine which tanks to show based on user profile assignment
    tanks = get_user_tanks(request.user).order_by("tank_id")
    flow_meters = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    tank_list = []
    for tank in tanks:
        latest = tank.readings.first()
        level = latest.level_percent if latest else 0.0
        tank_list.append({
            "id": tank.tank_id,
            "name": tank.name,
            "level": level,
        })
        
    flow_meter_list = []
    for fm in flow_meters:
        latest = fm.readings.first()
        flow = latest.flow_rate if latest else 0.0
        flow_meter_list.append({
            "id": fm.meter_id,
            "name": fm.name,
            "flow_rate": flow,
            "unit": fm.flow_unit
        })
    
    # Calculate alarm counts for sidebar badge
    high_alarms = Alarm.objects.filter(alarm_type="high", acknowledged=False).count()
    low_alarms = Alarm.objects.filter(alarm_type="low", acknowledged=False).count()

    context = {
        "tanks": tank_list,
        "flow_meters": flow_meter_list,
        "high_alarms": high_alarms,
        "low_alarms": low_alarms,
        "current_time": timezone.now(),
        "user": request.user,
        "is_connected": is_connected,
        "connection_msg": connection_msg,
    }
    return render(request, "core/trend_chart.html", context)


@login_required
@check_license_expiry
def alarms_view(request):
    """Page: Alarm Management."""
    # Get last Modbus daemon status log to verify active connection
    last_log = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    is_connected = False
    connection_msg = "Disconnected"
    if last_log:
        connection_msg = last_log.message
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True

    # Determine active tanks based on user profile assignment
    tanks = get_user_tanks(request.user).order_by("tank_id")

    from django.db.models import Q
    flow_meters = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    # Query real alarm records from database for both tanks and flow meters
    db_alarms = Alarm.objects.filter(
        Q(tank__in=tanks) | Q(flow_meter__in=flow_meters)
    ).select_related("tank", "flow_meter").order_by("-timestamp")
    
    # Calculate live stats
    high_alarms_count = db_alarms.filter(alarm_type="high", acknowledged=False).count()
    low_alarms_count = db_alarms.filter(alarm_type="low", acknowledged=False).count()
    acknowledged_count = db_alarms.filter(acknowledged=True).count()
    
    # Total active (unacknowledged) alarms
    active_alarms_count = high_alarms_count + low_alarms_count

    events = []
    for alarm in db_alarms:
        device_id = alarm.tank.tank_id if alarm.tank else (alarm.flow_meter.meter_id if alarm.flow_meter else "Unknown")
        device_name = alarm.tank.name if alarm.tank else (alarm.flow_meter.name if alarm.flow_meter else "Unknown")
        
        if alarm.tank:
            high_limit = format_tank_level(alarm.tank.high_limit, alarm.tank.capacity_kl, alarm.tank.unit, alarm.tank.high_limit)
            low_limit = format_tank_level(alarm.tank.low_limit, alarm.tank.capacity_kl, alarm.tank.unit, alarm.tank.low_limit)
            if alarm.tank.unit == "RAW":
                formatted_level = format_tank_level(None, alarm.tank.capacity_kl, alarm.tank.unit, alarm.level_percent)
            else:
                formatted_level = format_tank_level(alarm.level_percent, alarm.tank.capacity_kl, alarm.tank.unit)
            message = f"Value exceeded High Limit ({high_limit})" if alarm.alarm_type == "high" else f"Value dropped below Low Limit ({low_limit})"
        elif alarm.flow_meter:
            high_limit = f"{alarm.flow_meter.high_limit} {alarm.flow_meter.flow_unit}"
            low_limit = f"{alarm.flow_meter.low_limit} {alarm.flow_meter.flow_unit}"
            formatted_level = f"{alarm.level_percent:.1f} {alarm.flow_meter.flow_unit}"
            message = f"Flow rate exceeded High Limit ({high_limit})" if alarm.alarm_type == "high" else f"Flow rate dropped below Low Limit ({low_limit})"
        else:
            high_limit = low_limit = "0"
            formatted_level = f"{alarm.level_percent:.1f}"
            message = f"Value exceeded High Limit ({high_limit})" if alarm.alarm_type == "high" else f"Value dropped below Low Limit ({low_limit})"

        events.append({
            "id": alarm.pk,
            "time": timezone.localtime(alarm.timestamp).strftime("%d %b %Y %I:%M:%S %p"),
            "tank_id": device_id,
            "tank_name": device_name,
            "type": "High Alarm" if alarm.alarm_type == "high" else "Low Alarm",
            "severity": "HIGH" if alarm.alarm_type == "high" else "LOW",
            "level": alarm.level_percent,
            "formatted_level": formatted_level,
            "message": message,
            "status": "Acknowledged" if alarm.acknowledged else "Active",
        })

    # Compile alarm settings configuration per tank
    settings_list = []
    for tank in tanks:
        settings_list.append({
            "id": tank.tank_id,
            "name": tank.name,
            "high_limit": tank.high_limit,
            "low_limit": tank.low_limit,
            "actions": "Email, SMS, Buzzer",
            "status": "Enabled" if tank.is_active else "Disabled",
        })

    total_devices_count = tanks.count() + flow_meters.count()
    normal_count = total_devices_count - active_alarms_count
    if normal_count < 0:
        normal_count = 0

    context = {
        "high_alarms": high_alarms_count,
        "low_alarms": low_alarms_count,
        "acknowledged": acknowledged_count,
        "normal": normal_count,
        "total_alarms": db_alarms.count(),
        "events": events,
        "settings": settings_list,
        "current_time": timezone.now(),
        "user": request.user,
        "is_connected": is_connected,
        "connection_msg": connection_msg,
    }
    return render(request, "core/alarms.html", context)


@login_required
@check_license_expiry
def reports_view(request):
    """Page: Reports."""
    # Get last Modbus daemon status log to verify active connection
    last_log = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    is_connected = False
    if last_log:
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True

    # Determine active tanks based on user profile assignment
    all_tanks_list = get_user_tanks(request.user).order_by("tank_id")
    all_flow_meters_list = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    # Parse GET filters
    report_type = request.GET.get("report_type", "daily").lower()
    if report_type not in ["daily", "weekly", "monthly", "custom"]:
        report_type = "daily"

    device_id = request.GET.get("device_id", "all")
    if device_id.startswith("tank_"):
        t_id = device_id.split("_")[1]
        tanks = all_tanks_list.filter(tank_id=t_id)
        flow_meters = FlowMeter.objects.none()
    elif device_id.startswith("fm_"):
        fm_id = device_id.split("_")[1]
        tanks = Tank.objects.none()
        flow_meters = all_flow_meters_list.filter(meter_id=fm_id)
    else:
        tanks = all_tanks_list
        flow_meters = all_flow_meters_list

    # Parse base date
    date_str = request.GET.get("date", "").strip()
    base_date = None
    formats_to_try = [
        "%Y-%m-%d", "%Y_%m-%d", "%Y/%m/%d",
        "%d-%m-%Y", "%d_%m-%Y", "%d/%m-%Y",
        "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S"
    ]
    for fmt in formats_to_try:
        try:
            base_date = datetime.strptime(date_str, fmt)
            break
        except (ValueError, TypeError):
            continue
    if not base_date:
        base_date = timezone.now()
    
    date_value = base_date.strftime("%Y-%m-%d")

    # Parse end date for custom reports
    end_date_str = request.GET.get("end_date", "").strip()
    custom_end = None
    for fmt in formats_to_try:
        try:
            custom_end = datetime.strptime(end_date_str, fmt)
            break
        except (ValueError, TypeError):
            continue
    if not custom_end:
        custom_end = base_date

    end_date_value = custom_end.strftime("%Y-%m-%d")

    # Determine start and end ranges
    if report_type == "weekly":
        start_dt = base_date - timedelta(days=6)
        end_dt = base_date
    elif report_type == "monthly":
        start_dt = base_date - timedelta(days=29)
        end_dt = base_date
    elif report_type == "custom":
        start_dt = base_date
        end_dt = custom_end
    else:  # daily
        start_dt = base_date
        end_dt = base_date

    # Normalize to start and end of day
    start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Convert to timezone aware if required
    from django.conf import settings
    if settings.USE_TZ:
        if timezone.is_naive(start_dt):
            start_dt = timezone.make_aware(start_dt)
        if timezone.is_naive(end_dt):
            end_dt = timezone.make_aware(end_dt)

    from django.db.models import Avg, Min, Max
    
    summary_records = []
    fm_records = []
    
    total_open_sum = 0
    total_close_sum = 0
    tanks_count = 0
    primary_tank_unit = "RAW"
    
    fm_open_sum = 0
    fm_close_sum = 0
    fm_count = 0
    primary_fm_unit = "Liters"
    
    highest_tank_label = "--"
    lowest_tank_label = "--"
    total_high_alarms = 0
    total_low_alarms = 0
    total_fm_alarms = 0

    for tank in tanks:
        # Get historical readings of this tank for the selected range
        readings = TankReading.objects.filter(tank=tank, timestamp__range=(start_dt, end_dt))

        # Get alarms in the selected range
        t_high_alarms = Alarm.objects.filter(tank=tank, alarm_type="high", timestamp__range=(start_dt, end_dt)).count()
        t_low_alarms = Alarm.objects.filter(tank=tank, alarm_type="low", timestamp__range=(start_dt, end_dt)).count()

        if tank.widget_type == 'flow_meter':
            primary_fm_unit = tank.total_unit or "Liters"
            if readings.exists():
                oldest_reading = readings.order_by("timestamp").first()
                newest_reading = readings.order_by("-timestamp").first()
                
                opening_val = oldest_reading.total_flow if (oldest_reading and oldest_reading.total_flow is not None) else 0.0
                closing_val = newest_reading.total_flow if (newest_reading and newest_reading.total_flow is not None) else 0.0
                diff_val = closing_val - opening_val
                
                opening_disp_str = f"{round(opening_val, 2)} {tank.total_unit}"
                closing_disp_str = f"{round(closing_val, 2)} {tank.total_unit}"
                diff_disp = f"{round(diff_val, 2)} {tank.total_unit}"
                
                fm_open_sum += opening_val
                fm_close_sum += closing_val
                fm_count += 1
            else:
                opening_disp_str = "--"
                closing_disp_str = "--"
                diff_disp = "--"
                
            fm_records.append({
                "id": tank.tank_id,
                "name": tank.name,
                "opening": opening_disp_str,
                "closing": closing_disp_str,
                "difference": diff_disp,
                "total_unit": tank.total_unit
            })
        else:
            primary_tank_unit = tank.unit or "RAW"
            if readings.exists():
                oldest_reading = readings.order_by("timestamp").first()
                newest_reading = readings.order_by("-timestamp").first()
                
                opening_disp = format_tank_level(oldest_reading.level_percent, tank.capacity_kl, tank.unit, oldest_reading.raw_value) if oldest_reading else "--"
                closing_disp = format_tank_level(newest_reading.level_percent, tank.capacity_kl, tank.unit, newest_reading.raw_value) if newest_reading else "--"
                
                diff_pct = newest_reading.level_percent - oldest_reading.level_percent if (newest_reading and oldest_reading) else 0.0
                diff_raw = newest_reading.raw_value - oldest_reading.raw_value if (newest_reading and oldest_reading and newest_reading.raw_value is not None and oldest_reading.raw_value is not None) else None
                diff_disp = format_tank_level(diff_pct, tank.capacity_kl, tank.unit, diff_raw)
                
                if oldest_reading:
                    if tank.unit == "RAW":
                        open_v = oldest_reading.raw_value if oldest_reading.raw_value is not None else 0.0
                    else:
                        open_v = get_tank_value_in_unit(oldest_reading.level_percent, tank.capacity_kl, tank.unit, oldest_reading.raw_value)
                    total_open_sum += open_v
                if newest_reading:
                    if tank.unit == "RAW":
                        close_v = newest_reading.raw_value if newest_reading.raw_value is not None else 0.0
                    else:
                        close_v = get_tank_value_in_unit(newest_reading.level_percent, tank.capacity_kl, tank.unit, newest_reading.raw_value)
                    total_close_sum += close_v
                tanks_count += 1
            else:
                opening_disp = "--"
                closing_disp = "--"
                diff_disp = "--"

            summary_records.append({
                "id": tank.tank_id,
                "name": tank.name,
                "opening": opening_disp,
                "closing": closing_disp,
                "difference": diff_disp,
            })

        total_high_alarms += t_high_alarms
        total_low_alarms += t_low_alarms

    # Across all summary totals for Tanks
    tank_diff_sum = total_close_sum - total_open_sum
    tank_open_disp = format_total_val(total_open_sum, primary_tank_unit)
    tank_close_disp = format_total_val(total_close_sum, primary_tank_unit)
    tank_diff_disp = format_total_val(tank_diff_sum, primary_tank_unit)
    total_capacity = sum([t.capacity_kl for t in tanks if t.widget_type != 'flow_meter'])

    # Generate trend chart labels and datasets dynamically based on DB values
    chart_labels = []
    chart_datasets = []
    
    if tanks.exists():
        if report_type == "daily":
            # 2-hour intervals for a single day
            chart_labels = [f"{h:02d}:00" for h in range(0, 25, 2)]
            colors = ['#4CAF50', '#2196F3', '#FF9800', '#9C27B0', '#00BCD4']
            for index, tank in enumerate(tanks[:5]): # limit to top 5 for visual clarity
                data_points = []
                for h in range(0, 25, 2):
                    if h == 24:
                        h_start = (start_dt + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
                        h_end = h_start
                    else:
                        h_start = start_dt.replace(hour=h, minute=0, second=0, microsecond=0)
                        h_end = h_start.replace(hour=h+1, minute=59, second=59, microsecond=999999)
                    avg_l = TankReading.objects.filter(tank=tank, timestamp__range=(h_start, h_end)).aggregate(avg=Avg("level_percent"))["avg"]
                    data_points.append(round(avg_l, 1) if avg_l is not None else 0.0)
                chart_datasets.append({
                    "label": tank.name,
                    "data": data_points,
                    "borderColor": colors[index % len(colors)],
                    "borderWidth": 1.5,
                    "pointRadius": 0,
                    "tension": 0.35,
                    "fill": False
                })
        else:
            # Weekly/Monthly/Custom - group by day
            delta = end_dt - start_dt
            days_count = max(1, delta.days + 1)
            chart_labels = []
            for i in range(days_count):
                d = start_dt + timedelta(days=i)
                chart_labels.append(d.strftime("%d %b"))
                
            colors = ['#4CAF50', '#2196F3', '#FF9800', '#9C27B0', '#00BCD4']
            for index, tank in enumerate(tanks[:5]):
                data_points = []
                for i in range(days_count):
                    d_start = (start_dt + timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
                    d_end = d_start.replace(hour=23, minute=59, second=59, microsecond=999999)
                    avg_l = TankReading.objects.filter(tank=tank, timestamp__range=(d_start, d_end)).aggregate(avg=Avg("level_percent"))["avg"]
                    data_points.append(round(avg_l, 1) if avg_l is not None else 0.0)
                chart_datasets.append({
                    "label": tank.name,
                    "data": data_points,
                    "borderColor": colors[index % len(colors)],
                    "borderWidth": 1.5,
                    "pointRadius": 0,
                    "tension": 0.35,
                    "fill": False
                })

    # Generate Flow Meter summary records for standalone FlowMeters
    for fm in flow_meters:
        unit_str = fm.total_unit or fm.flow_unit or "m³/h"
        primary_fm_unit = unit_str
        readings = FlowMeterReading.objects.filter(flow_meter=fm, timestamp__range=(start_dt, end_dt))
        
        f_high_alarms = Alarm.objects.filter(flow_meter=fm, alarm_type="high", timestamp__range=(start_dt, end_dt)).count()
        f_low_alarms = Alarm.objects.filter(flow_meter=fm, alarm_type="low", timestamp__range=(start_dt, end_dt)).count()
        
        if readings.exists():
            oldest = readings.order_by("timestamp").first()
            newest = readings.order_by("-timestamp").first()
            opening_val = round(oldest.total_volume, 2) if oldest else 0.0
            closing_val = round(newest.total_volume, 2) if newest else 0.0
            diff_val = closing_val - opening_val
            
            opening_disp_str = f"{round(opening_val, 2)} {unit_str}"
            closing_disp_str = f"{round(closing_val, 2)} {unit_str}"
            diff_disp = f"{round(diff_val, 2)} {unit_str}"
            
            fm_open_sum += opening_val
            fm_close_sum += closing_val
            fm_count += 1
        else:
            opening_disp_str = "--"
            closing_disp_str = "--"
            diff_disp = "--"

        total_fm_alarms += (f_high_alarms + f_low_alarms)

        fm_records.append({
            "id": fm.meter_id,
            "name": fm.name,
            "opening": opening_disp_str,
            "closing": closing_disp_str,
            "difference": diff_disp,
            "total_unit": unit_str
        })

    # Across all summary totals for Flow Meters
    fm_diff_sum = fm_close_sum - fm_open_sum
    fm_open_disp = f"{round(fm_open_sum, 2)} {primary_fm_unit}"
    fm_close_disp = f"{round(fm_close_sum, 2)} {primary_fm_unit}"
    fm_diff_disp = f"{round(fm_diff_sum, 2)} {primary_fm_unit}"

    # Fetch chronological data changes
    changes = get_data_changes(tanks, flow_meters, start_dt, end_dt)

    import json
    context = {
        "report_type": report_type,
        "date_value": date_value,
        "end_date_value": end_date_value,
        "device_id": device_id,
        "all_tanks_list": all_tanks_list,
        "all_flow_meters_list": all_flow_meters_list,
        "tanks": tanks,
        "records": summary_records,
        "fm_records": fm_records,
        "total_tanks": tanks.count(),
        "total_capacity": round(total_capacity, 1),
        "tank_open_disp": tank_open_disp,
        "tank_close_disp": tank_close_disp,
        "tank_diff_disp": tank_diff_disp,
        "tank_diff_val": tank_diff_sum,
        "fm_open_disp": fm_open_disp,
        "fm_close_disp": fm_close_disp,
        "fm_diff_disp": fm_diff_disp,
        "fm_diff_val": fm_diff_sum,
        "lowest_level": lowest_tank_label,
        "highest_level": highest_tank_label,
        "high_alarms": total_high_alarms,
        "low_alarms": total_low_alarms,
        "total_alarms": total_high_alarms + total_low_alarms + total_fm_alarms,
        "chart_labels": json.dumps(chart_labels),
        "chart_datasets": json.dumps(chart_datasets),
        "changes": changes,
        "current_time": timezone.now(),
    }
    return render(request, "core/reports.html", context)


@login_required
@check_license_expiry
def data_log_view(request):
    """Page: Data Log."""
    # Get last Modbus daemon status log to verify active connection
    last_log = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    is_connected = False
    if last_log:
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True

    # Determine active tanks based on user profile assignment
    tanks = get_user_tanks(request.user).order_by("tank_id")
    flow_meters = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    # Query real level history logs from database
    db_readings = list(TankReading.objects.filter(tank__in=tanks).select_related("tank").order_by("-timestamp")[:250])
    db_fm_readings = list(FlowMeterReading.objects.filter(flow_meter__in=flow_meters).select_related("flow_meter").order_by("-timestamp")[:250])
    
    # Compile readings into time logs matrix rows
    time_groups = {}
    for r in db_readings:
        time_str = timezone.localtime(r.timestamp).strftime("%d %b %Y %I:%M %p")
        if time_str not in time_groups:
            time_groups[time_str] = {}
        if r.tank.widget_type == 'flow_meter':
            time_groups[time_str][f"tank_{r.tank.tank_id}_rate"] = f"{r.flow_rate} {r.tank.flow_unit}" if r.flow_rate is not None else "--"
            time_groups[time_str][f"tank_{r.tank.tank_id}_total"] = f"{r.total_flow} {r.tank.total_unit}" if r.total_flow is not None else "--"
        else:
            time_groups[time_str][r.tank.tank_id] = format_tank_level(r.level_percent, r.tank.capacity_kl, r.tank.unit, getattr(r, 'raw_value', None))
        
    for r in db_fm_readings:
        time_str = timezone.localtime(r.timestamp).strftime("%d %b %Y %I:%M %p")
        if time_str not in time_groups:
            time_groups[time_str] = {}
        time_groups[time_str][f"fm_{r.flow_meter.meter_id}_rate"] = f"{r.flow_rate} {r.flow_meter.flow_unit}" if r.flow_rate is not None else "--"
        time_groups[time_str][f"fm_{r.flow_meter.meter_id}_total"] = f"{r.total_volume} {r.flow_meter.total_unit}" if r.total_volume is not None else "--"

    # Sort timestamps in reverse chronological order
    sorted_times = sorted(time_groups.keys(), reverse=True)

    matrix_rows = []
    for time_str in sorted_times:
        vals = time_groups[time_str]
        row_values = []
        for tank in tanks:
            if tank.widget_type == 'flow_meter':
                rate = vals.get(f"tank_{tank.tank_id}_rate", "--")
                total = vals.get(f"tank_{tank.tank_id}_total", "--")
                row_values.append(rate)
                row_values.append(total)
            else:
                val = vals.get(tank.tank_id, "--")
                row_values.append(val)
        for fm in flow_meters:
            rate = vals.get(f"fm_{fm.meter_id}_rate", "--")
            total = vals.get(f"fm_{fm.meter_id}_total", "--")
            row_values.append(rate)
            row_values.append(total)
        matrix_rows.append({
            "time": time_str,
            "values": row_values
        })

    # Badge alarm counts
    high_alarms = Alarm.objects.filter(alarm_type="high", acknowledged=False).count()
    low_alarms = Alarm.objects.filter(alarm_type="low", acknowledged=False).count()

    # Get boundary date timestamps
    from_date = None
    to_date = None
    if sorted_times:
        to_date = sorted_times[0]
        from_date = sorted_times[-1]

    context = {
        "rows": matrix_rows,
        "active_tanks": tanks,
        "active_flow_meters": flow_meters,
        "total_records": len(sorted_times),
        "from_date": from_date,
        "to_date": to_date,
        "high_alarms": high_alarms,
        "low_alarms": low_alarms,
        "current_time": timezone.now(),
        "user": request.user,
        "is_connected": is_connected,
    }
    return render(request, "core/data_log.html", context)


def get_available_com_ports(current_port=None):
    import serial.tools.list_ports
    # Get only physical ports from the OS
    ports = [p.device for p in serial.tools.list_ports.comports()]
    ports = sorted(list(set(ports)))
    # Always append SIMULATOR at the very end
    if "SIMULATOR" not in ports:
        ports.append("SIMULATOR")
    return ports


def _is_profile_connected(profile_name):
    """Check if a specific profile's Modbus daemon is actively communicating."""
    last_log = EventLog.objects.filter(source=f"Modbus Daemon ({profile_name})").order_by("-timestamp").first()
    if last_log:
        msg = last_log.message
        if "Successfully read" in msg or "Telemetry scan successful" in msg:
            if "unavailable" not in msg and "failed" not in msg:
                return True, msg
    # Fallback to generic source
    last_log2 = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    if last_log2:
        msg2 = last_log2.message
        if "Successfully read" in msg2 or "Telemetry scan successful" in msg2:
            if "unavailable" not in msg2 and "failed" not in msg2:
                return True, msg2
    return False, "Disconnected"


@login_required
@check_license_expiry
def settings_view(request):
    """Page: Settings."""
    # Build list of mapping registers from database tanks dynamically
    tanks = Tank.objects.all().order_by("tank_id")
    registers = []
    for i, tank in enumerate(tanks, start=1):
        registers.append({
            "no": i,
            "id": tank.tank_id,
            "name": tank.name,
            "address": tank.register_address,
            "type": tank.data_type,
            "scale": "Linear",
            "min": 0,
            "max": 100,
            "unit": tank.unit,
            "widget_type": tank.widget_type,
            "capacity": tank.capacity_kl,
            "capacity_unit": getattr(tank, 'capacity_unit', 'KL') or 'KL',
            "high_limit": tank.high_limit,
            "low_limit": tank.low_limit,
            "alarm_enabled": tank.alarm_enabled,
            "raw_zero": tank.raw_zero,
            "raw_span": tank.raw_span,
            "scanner_raw_zero": tank.scanner_raw_zero,
            "scanner_raw_span": tank.scanner_raw_span,
            "slave_id": tank.slave_id,
            "function_code": tank.function_code,
            "byte_order": tank.byte_order,
            "scaling": tank.scaling,
            "offset": tank.offset,
            "flow_rate_register": tank.flow_rate_register,
            "total_volume_register": tank.total_volume_register,
            "flow_unit": tank.flow_unit,
            "total_unit": tank.total_unit,
            "error_accuracy": tank.error_accuracy,
        })

    # Auto-initialize SimulatedRegister entries for all mapped registers
    from .models import SimulatedRegister, FlowMeter
    for tank in tanks:
        if tank.widget_type == 'flow_meter':
            if tank.flow_rate_register is not None:
                SimulatedRegister.objects.get_or_create(
                    register_address=tank.flow_rate_register,
                    defaults={"value": 15}
                )
            if tank.total_volume_register is not None:
                SimulatedRegister.objects.get_or_create(
                    register_address=tank.total_volume_register,
                    defaults={"value": 120}
                )
        else:
            if tank.register_address is not None:
                SimulatedRegister.objects.get_or_create(
                    register_address=tank.register_address,
                    defaults={"value": 50}
                )
    for fm in FlowMeter.objects.all():
        if fm.flow_rate_register is not None:
            SimulatedRegister.objects.get_or_create(
                register_address=fm.flow_rate_register,
                defaults={"value": 15}
            )
        if fm.total_volume_register is not None:
            SimulatedRegister.objects.get_or_create(
                register_address=fm.total_volume_register,
                defaults={"value": 120}
            )
    sim_registers = list(SimulatedRegister.objects.all().order_by("register_address"))

    # Build list of mapping registers for flow meters
    flow_meters_qs = FlowMeter.objects.all().order_by("meter_id")
    flow_registers = []
    for fm in flow_meters_qs:
        flow_registers.append({
            "id": fm.meter_id,
            "name": fm.name,
            "flow_rate_register": fm.flow_rate_register,
            "total_volume_register": fm.total_volume_register,
            "slave_id": fm.slave_id,
            "function_code": fm.function_code,
            "data_type": fm.data_type,
            "byte_order": fm.byte_order,
            "flow_unit": fm.flow_unit,
            "total_unit": fm.total_unit,
            "high_limit": fm.high_limit,
            "low_limit": fm.low_limit,
            "alarm_enabled": fm.alarm_enabled,
            "error_accuracy": fm.error_accuracy,
            "scanner_raw_zero": fm.scanner_raw_zero,
            "scanner_raw_span": fm.scanner_raw_span,
            "calibrated_span": fm.calibrated_span,
        })

    # Badge alarm counts
    high_alarms = Alarm.objects.filter(alarm_type="high", acknowledged=False).count()
    low_alarms = Alarm.objects.filter(alarm_type="low", acknowledged=False).count()

    # Get last Modbus daemon status log
    last_log = EventLog.objects.filter(source="Modbus Daemon").order_by("-timestamp").first()
    is_connected = False
    connection_msg = "Disconnected"
    if last_log:
        connection_msg = last_log.message
        if "Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message:
            if "unavailable" not in last_log.message and "failed" not in last_log.message:
                is_connected = True
    # Also check profile-specific daemon logs
    if not is_connected:
        tank_log = EventLog.objects.filter(source="Modbus Daemon (Tanks)").order_by("-timestamp").first()
        if tank_log and ("Successfully read" in tank_log.message or "Telemetry scan successful" in tank_log.message):
            if "unavailable" not in tank_log.message and "failed" not in tank_log.message:
                is_connected = True
                connection_msg = tank_log.message

    # Get Serial Connection Configs
    from .models import SerialConnectionConfig
    tank_config, _ = SerialConnectionConfig.objects.get_or_create(
        profile_name="Tanks",
        defaults={
            "connection_name": "Masibus Scanner 01",
            "device_type": "Masibus Scanner",
            "connection_type": "Serial (RS485)",
            "com_port": "COM3",
            "baud_rate": 9600,
            "data_bits": 8,
            "parity": "None",
            "stop_bits": 1,
            "slave_id": 1
        }
    )

    flow_config, _ = SerialConnectionConfig.objects.get_or_create(
        profile_name="Flow Meters",
        defaults={
            "connection_name": "Flow Meter Scanner 01",
            "device_type": "Modbus RS485",
            "connection_type": "Serial (RS485)",
            "com_port": "COM4",
            "baud_rate": 9600,
            "data_bits": 8,
            "parity": "None",
            "stop_bits": 1,
            "slave_id": 1
        }
    )

    # On initial page load, show the currently saved COM port, physically available ports, and SIMULATOR
    available_ports = get_available_com_ports(tank_config.com_port)
    available_ports_flow = get_available_com_ports(flow_config.com_port)

    is_tank_connected, tank_conn_msg = _is_profile_connected("Tanks")
    is_flow_connected, flow_conn_msg = _is_profile_connected("Flow Meters")
    is_connected = is_tank_connected or is_flow_connected
    connection_msg = tank_conn_msg if is_tank_connected else flow_conn_msg

    from .models import LicenseConfig, LicenseKey, WorkspaceConfig
    from django.contrib.auth.models import User
    license_config = LicenseConfig.load()
    workspace_config = WorkspaceConfig.objects.first()
    
    display_license_type = "Enterprise License"
    activation_key = ""
    if workspace_config:
        activation_key = workspace_config.activation_key or ""
        key_str = activation_key.strip()
        try:
            lic = LicenseKey.objects.get(key=key_str)
            if lic.duration_days == 730:
                display_license_type = "2-Year Valid License"
            elif lic.duration_days == 1095:
                display_license_type = "3-Year Valid License"
            elif lic.duration_days >= 36500:
                display_license_type = "Lifetime License"
            else:
                display_license_type = "1-Year Valid License"
        except LicenseKey.DoesNotExist:
            if key_str == 'RADIOGEET-AXIONIX-S3CR3T-K3Y-2026':
                display_license_type = "Lifetime License"
            elif key_str == 'RADIOGEET-AXIONIX-S3CR3T-K3Y-2026-3T':
                display_license_type = "3-Day Trial License"
            elif key_str.startswith('RADIOGEET-AXIONIX-KEY-2026-'):
                display_license_type = "1-Year Valid License"
            else:
                display_license_type = "Enterprise License"

    masked_license_key = "No License Key"
    if activation_key:
        if activation_key.upper().startswith("RADIOGEET-"):
            suffix = activation_key[10:]
            masked_suffix = "".join(['X' if c.isalnum() else c for c in suffix])
            masked_license_key = "RADIOGEET-" + masked_suffix
        else:
            masked_license_key = activation_key[:9] + "".join(['X' if c.isalnum() else c for c in activation_key[9:]])

    signup_user = User.objects.exclude(username='admin').order_by('id').first()
    if not signup_user:
        signup_user = User.objects.order_by('id').first()
    activated_for_username = signup_user.username if signup_user else "N/A"

    context = {
        "display_license_type": display_license_type,
        "masked_license_key": masked_license_key,
        "activated_for_username": activated_for_username,
        "license_config": license_config,
        "registers": registers,
        "high_alarms": high_alarms,
        "low_alarms": low_alarms,
        "current_time": timezone.now(),
        "user": request.user,
        "is_connected": is_connected,
        "is_tank_connected": is_tank_connected,
        "is_flow_connected": is_flow_connected,
        "connection_msg": connection_msg,
        "tank_config": tank_config,
        "flow_config": flow_config,
        "available_ports": available_ports,
        "available_ports_flow": available_ports_flow,
        "sim_registers": sim_registers,
        "flow_registers": flow_registers,
    }
    return render(request, "core/settings.html", context)


@login_required
def settings_list_com_ports_view(request):
    try:
        # Scan and return available ports including current port and SIMULATOR option
        from .models import SerialConnectionConfig
        tank_config = SerialConnectionConfig.objects.filter(profile_name="Tanks").first()
        flow_config = SerialConnectionConfig.objects.filter(profile_name="Flow Meters").first()
        ports = get_available_com_ports()
        # ports already contains physical ports + SIMULATOR
        return JsonResponse({"success": True, "ports": ports, "ports_flow": ports})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
def settings_save_connection_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"success": False, "error": "Invalid request body."})
        
        from .models import SerialConnectionConfig
        profile_name = data.get("profile_name", "Tanks")
        config, _ = SerialConnectionConfig.objects.get_or_create(profile_name=profile_name)
            
        config.connection_name = data.get("connection_name", "Masibus Scanner 01")
        config.device_type = data.get("device_type", "Masibus Scanner")
        config.connection_type = data.get("connection_type", "Serial (RS485)")
        config.com_port = data.get("com_port", "COM3")
        try:
            config.baud_rate = int(data.get("baud_rate", 9600))
            config.data_bits = int(data.get("data_bits", 8))
            config.parity = data.get("parity", "None")
            config.stop_bits = int(data.get("stop_bits", 1))
            config.slave_id = int(data.get("slave_id", 1))
            
            config.timeout = float(data.get("timeout", 0.5))
            config.retry_count = int(data.get("retry_count", 3))
            config.polling_interval = float(data.get("polling_interval", 5.0))
            config.response_delay = float(data.get("response_delay", 0.0))
            config.rts_delay = float(data.get("rts_delay", 0.0))
            config.auto_reconnect = bool(data.get("auto_reconnect", True))
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "error": "Baud rate, data bits, stop bits, slave ID, timeout, and numeric settings must be valid numbers."})
            
        config.save()
        
        try:
            from .modbus_daemon import clear_device_snooze
            clear_device_snooze()
        except Exception:
            pass
        
        # Log system event
        from .models import EventLog
        EventLog.objects.create(
            event_type="SYSTEM",
            source="Settings",
            message=f"Saved Serial Connection Config for {profile_name}: Port={config.com_port}, Baud={config.baud_rate}"
        )
        
        return JsonResponse({"success": True, "message": "Settings saved successfully."})
        
    return JsonResponse({"success": False, "error": "Method not allowed."})


@login_required
def settings_test_connection_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"success": False, "error": "Invalid request body."})
        
        try:
            from .modbus_daemon import clear_device_snooze
            clear_device_snooze()
        except Exception:
            pass
            
        com_port = data.get("com_port", "COM3")

        if com_port == "SIMULATOR":
            from .models import SimulatedRegister
            val_obj, _ = SimulatedRegister.objects.get_or_create(register_address=0, defaults={"value": 50})
            return JsonResponse({
                "success": True,
                "message": f"Connection to SIMULATOR successful! Database simulation active (Register 0 = {val_obj.value})."
            })
        
        # Check if background daemon is already running successfully on the same port
        from .models import EventLog
        profile_name = data.get("profile_name", "Tanks")
        last_log = EventLog.objects.filter(source=f"Modbus Daemon ({profile_name})").order_by("-timestamp").first()
        if not last_log:
            last_log = EventLog.objects.filter(source__icontains="Modbus Daemon").order_by("-timestamp").first()

        if last_log and ("Successfully read" in last_log.message or "Telemetry scan successful" in last_log.message) and f"on {com_port}" in last_log.message:
            return JsonResponse({
                "success": True,
                "message": f"Connection to {com_port} is already active and communicating successfully via background service."
            })
        try:
            baudrate = int(data.get("baud_rate", 9600))
            bytesize = int(data.get("data_bits", 8))
            parity_str = data.get("parity", "None")
            stopbits = int(data.get("stop_bits", 1))
            slave_id = int(data.get("slave_id", 1))
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "error": "Parameters must be valid numbers."})
            
        import serial
        import serial.tools.list_ports
        try:
            import minimalmodbus
            
            parity_map = {
                "None": serial.PARITY_NONE,
                "Even": serial.PARITY_EVEN,
                "Odd": serial.PARITY_ODD,
            }
            parity = parity_map.get(parity_str, serial.PARITY_NONE)
            
            # RadioDAQ pattern: check if port is physically available first
            available_ports = [p.device for p in serial.tools.list_ports.comports()]
            if com_port not in available_ports:
                return JsonResponse({
                    "success": False,
                    "error": f"Port {com_port} not available. Device might be disconnected. Available ports: {', '.join(available_ports) if available_ports else 'None'}"
                })
            
            # Replaced redundant serial check with single robust minimalmodbus connection
            from .modbus_daemon import get_com_lock, decode_registers
            with get_com_lock(com_port):
                instrument = None
                try:
                    instrument = minimalmodbus.Instrument(com_port, slave_id, close_port_after_each_call=True)
                    instrument.serial.baudrate = baudrate
                    instrument.serial.bytesize = bytesize
                    instrument.serial.parity = parity
                    instrument.serial.stopbits = stopbits
                    instrument.serial.timeout = 1.5
                    instrument.clear_buffers_before_each_transaction = True
                    
                    profile_name = data.get("profile_name", "Tanks")
                    test_success = False
                    test_val = None
                    test_info = ""
                    
                    # Strategy 1: Read from actual configured device register
                    if profile_name == "Tanks":
                        test_device = Tank.objects.filter(is_active=True).first()
                    else:
                        test_device = FlowMeter.objects.filter(is_active=True).first()
                    
                    if test_device:
                        raw_addr = test_device.register_address if hasattr(test_device, 'register_address') else test_device.flow_rate_register
                        # Auto-detect function code based on register range:
                        # 4XXXX Holding (FC3), 3XXXX Input (FC4)
                        if raw_addr >= 40000:
                            fc = 3
                            protocol_addr = raw_addr - 40001
                        elif raw_addr >= 30000:
                            fc = 4
                            protocol_addr = raw_addr - 30001
                        else:
                            fc = test_device.function_code
                            protocol_addr = raw_addr
                            
                        # Determine number of registers
                        num_regs = 1
                        data_type = getattr(test_device, 'data_type', 'UInt16')
                        if data_type in ("Int32", "UInt32", "Float32"):
                            num_regs = 2
                        elif data_type == "Float64":
                            num_regs = 4
                            
                        try:
                            if fc in (1, 2):
                                test_val = instrument.read_bit(protocol_addr, functioncode=fc)
                            else:
                                regs = instrument.read_registers(protocol_addr, num_regs, functioncode=fc)
                                test_val = decode_registers(regs, data_type, getattr(test_device, 'byte_order', 'ABCD'))
                            
                            test_success = True
                            test_info = f"Register {raw_addr} (FC{fc}) = {test_val}"
                        except Exception:
                            pass
                    
                    # Strategy 2: Try register 0 & 1 with FC4 (Input Registers - Masibus Scanner standard)
                    if not test_success:
                        for addr in [0, 1]:
                            # Try 1 register (UInt16/Int16) first for Masibus Scanners
                            try:
                                test_val = instrument.read_register(addr, functioncode=4)
                                test_success = True
                                test_info = f"Register {30001 if addr==0 else 30002} (FC4) = {test_val}"
                                break
                            except Exception:
                                pass
                            # Fallback to 2 registers (Float32/UInt32)
                            try:
                                regs = instrument.read_registers(addr, 2, functioncode=4)
                                test_val = decode_registers(regs, "Float32", "ABCD")
                                test_success = True
                                test_info = f"Register {30001 if addr==0 else 30002} (FC4, 32-bit) = {test_val}"
                                break
                            except Exception:
                                pass

                    # Strategy 3: Try register 0 & 1 with FC3 (Holding Registers)
                    if not test_success:
                        for addr in [0, 1]:
                            # Try 1 register (UInt16/Int16) first
                            try:
                                test_val = instrument.read_register(addr, functioncode=3)
                                test_success = True
                                test_info = f"Register {40001 if addr==0 else 40002} (FC3) = {test_val}"
                                break
                            except Exception:
                                pass
                            # Fallback to 2 registers (Float32/UInt32)
                            try:
                                regs = instrument.read_registers(addr, 2, functioncode=3)
                                test_val = decode_registers(regs, "Float32", "ABCD")
                                test_success = True
                                test_info = f"Register {40001 if addr==0 else 40002} (FC3, 32-bit) = {test_val}"
                                break
                            except Exception:
                                pass
                    
                    if test_success:
                        # Sanitize tiny exponent floats (< 1e-10) in test_info display
                        if isinstance(test_val, float) and abs(test_val) < 1e-10:
                            test_info = test_info.replace(str(test_val), "0")

                        # Log success so UI instantly shows 'Connected'
                        from .models import EventLog
                        EventLog.objects.create(
                            event_type="SYSTEM",
                            source="Modbus Daemon",
                            message=f"Telemetry scan successful. Connected manually on {com_port}."
                        )
                        
                        return JsonResponse({
                            "success": True,
                            "message": f"Connection to {com_port} successful! Device responded ({test_info})."
                        })
                    else:
                        return JsonResponse({
                            "success": True,
                            "warning": True,
                            "message": f"COM Port {com_port} opened successfully, but device did not respond. Tried Slave ID {slave_id} on multiple function codes. Check wiring, baud rate, and Slave ID."
                        })
                except serial.SerialException as ser_err:
                    err_str = str(ser_err)
                    if "PermissionError" in err_str or "Access is denied" in err_str:
                        return JsonResponse({
                            "success": True,
                            "message": f"Connection to {com_port} is active and communicating successfully via background service."
                        })
                    return JsonResponse({
                        "success": False,
                        "error": f"Failed to open {com_port}: {err_str}. Please check if port is in use or disconnected."
                    })
                except Exception as modbus_err:
                    return JsonResponse({
                        "success": True,
                        "warning": True,
                        "message": f"COM Port {com_port} opened successfully, but device did not respond. Modbus Error: {str(modbus_err)}. Check wiring, baud rate, and Slave ID."
                    })
                finally:
                    if instrument and instrument.serial and instrument.serial.is_open:
                        try:
                            instrument.serial.close()
                        except:
                            pass
        except Exception as e:
            return JsonResponse({"success": False, "error": f"Unexpected error during test: {str(e)}"})
            
    return JsonResponse({"success": False, "error": "Method not allowed."})

def settings_add_flow_meter_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
            meter_id = data.get("meter_id", "").strip()
            name = data.get("name", "").strip()
            if not meter_id or not name:
                return JsonResponse({"success": False, "error": "Meter ID and Name are required."})
                
            from .models import SerialConnectionConfig
            config = SerialConnectionConfig.objects.filter(profile_name="Flow Meters").first()
            if not config:
                config = SerialConnectionConfig.objects.create(profile_name="Flow Meters")
                
            com_port_param = data.get("com_port")
            if com_port_param:
                config.com_port = com_port_param.strip()
                config.save()
                
            from .models import FlowMeter, EventLog
            fm, created = FlowMeter.objects.get_or_create(meter_id=meter_id)
            fm.name = name
            
            # Map registers
            fm.flow_rate_register = int(data.get("flow_rate_register", 40001))
            
            total_vol_reg_val = data.get("total_volume_register")
            if total_vol_reg_val and str(total_vol_reg_val).strip():
                fm.total_volume_register = int(total_vol_reg_val)
            else:
                fm.total_volume_register = None
                
            fm.slave_id = int(data.get("slave_id", 1))
            fm.function_code = int(data.get("function_code", 3))
            fm.data_type = data.get("data_type", "Float32")
            fm.byte_order = data.get("byte_order", "ABCD")
            fm.flow_unit = data.get("flow_unit", "m³/h").strip() or "m³/h"
            
            total_unit_val = data.get("total_unit")
            if total_unit_val and str(total_unit_val).strip():
                fm.total_unit = str(total_unit_val).strip()
            else:
                fm.total_unit = fm.flow_unit
                
            def get_float(val, default):
                if val in (None, ""):
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default

            fm.high_limit = get_float(data.get("high_limit"), 90.0)
            fm.low_limit = get_float(data.get("low_limit"), 10.0)
            fm.error_accuracy = get_float(data.get("error_accuracy"), 0.0)
            fm.scanner_raw_zero = get_float(data.get("scanner_raw_zero"), 0.0)
            fm.scanner_raw_span = get_float(data.get("scanner_raw_span"), 0.0)
            fm.calibrated_span = get_float(data.get("calibrated_span"), 100.0)
            
            alarm_enabled_val = data.get("alarm_enabled", True)
            if isinstance(alarm_enabled_val, str):
                fm.alarm_enabled = alarm_enabled_val.lower() == 'true'
            else:
                fm.alarm_enabled = bool(alarm_enabled_val)
            
            fm.save()

            # Immediately create/re-calculate FlowMeterReading
            from .models import SimulatedRegister, FlowMeterReading
            sim_reg = SimulatedRegister.objects.filter(register_address=fm.flow_rate_register).first()
            if sim_reg is not None:
                raw_flow = float(sim_reg.value)
                if fm.scanner_raw_span > fm.scanner_raw_zero:
                    flow_rate = ((raw_flow - fm.scanner_raw_zero) / (fm.scanner_raw_span - fm.scanner_raw_zero)) * fm.calibrated_span
                    flow_rate = flow_rate + fm.error_accuracy
                else:
                    flow_rate = raw_flow + fm.error_accuracy
                FlowMeterReading.objects.create(
                    flow_meter=fm,
                    flow_rate=max(0.0, flow_rate),
                    total_volume=0.0
                )
            
            EventLog.objects.create(
                event_type="SYSTEM",
                source="Settings",
                message=f"Flow Meter {'Added' if created else 'Updated'}: {meter_id} ({name})"
            )
            
            return JsonResponse({"success": True, "message": f"Flow Meter '{name}' saved successfully."})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
            
    return JsonResponse({"success": False, "error": "Invalid request."})

def settings_delete_flow_meter_view(request, meter_id):
    if request.method == "POST":
        try:
            from .models import FlowMeter, EventLog
            fm = FlowMeter.objects.get(meter_id=meter_id)
            name = fm.name
            fm.delete()
            
            EventLog.objects.create(
                event_type="SYSTEM",
                source="Settings",
                message=f"Flow Meter Deleted: {meter_id} ({name})"
            )
            return JsonResponse({"success": True, "message": f"Flow Meter '{meter_id}' deleted successfully."})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request."})


def _create_demo_data():
    """Create 25 demo tanks with sample readings and alarms."""
    demo_levels = [
        78, 65, 42, 80, 56, 70, 33, 90, 48, 62,
        59, 84, 66, 27, 71, 58, 45, 92, 41, 30,
        57, 73, 68, 61, 50,
    ]

    now = timezone.now()

    for i in range(1, 26):
        tank = Tank.objects.create(
            tank_id=f"TANK {i}",
            name=f"Tank {i}",
            capacity_kl=5.0,
            location=f"Area {(i - 1) // 5 + 1}",
            is_active=True,
        )
        level = demo_levels[i - 1]

        # Create current reading
        TankReading.objects.create(tank=tank, level_percent=level)

        # Create historical readings for charts
        for h in range(24):
            historical_level = max(5, min(98, level + random.randint(-15, 15)))
            reading = TankReading(
                tank=tank,
                level_percent=historical_level,
            )
            reading.save()
            # Manually set timestamp for historical data
            TankReading.objects.filter(pk=reading.pk).update(
                timestamp=now - timedelta(hours=h)
            )

        # Create alarms for high/low levels
        if level >= 85:
            Alarm.objects.create(
                tank=tank,
                alarm_type="high",
                level_percent=level,
            )
        elif level <= 30:
            Alarm.objects.create(
                tank=tank,
                alarm_type="low",
                level_percent=level,
            )


@login_required
@check_license_expiry
def events_view(request):
    """Page: Events Log Page."""
    from .models import EventLog
    
    # Create demo events if none exist
    if EventLog.objects.count() == 0:
        EventLog.objects.create(
            event_type="SYSTEM",
            source="Connection",
            message="Masibus Scanner RS485 Modbus TCP Gateway registered successfully."
        )
        EventLog.objects.create(
            event_type="INFO",
            source="Modbus Daemon",
            message="Telemetry scan started: 25 tanks configured on COM3."
        )
        EventLog.objects.create(
            event_type="WARNING",
            source="Level Sensor",
            message="TANK 14 level dropped below low limit warning threshold."
        )
        EventLog.objects.create(
            event_type="CRITICAL",
            source="Level Sensor",
            message="TANK 18 critical High Alarm triggered: level exceeded 90% threshold."
        )

    logs = EventLog.objects.all().order_by("-timestamp")
    
    high_alarms = Alarm.objects.filter(alarm_type="high", acknowledged=False).count()
    low_alarms = Alarm.objects.filter(alarm_type="low", acknowledged=False).count()

    context = {
        "logs": logs,
        "high_alarms": high_alarms,
        "low_alarms": low_alarms,
        "current_time": timezone.now(),
        "user": request.user,
    }
    return render(request, "core/events.html", context)


def generate_pdf_response(headers, rows, title, filename, metadata=None, change_headers=None, change_rows=None, tank_rows=None, fm_rows=None):
    import os
    from django.conf import settings
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Setup document with professional margins
    doc = SimpleDocTemplate(
        response, 
        pagesize=letter,
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )
    elements = []
    styles = getSampleStyleSheet()
    
    # Dynamic styles based on column count and row count
    num_cols = len(headers)
    num_rows = len(rows)
    if num_cols > 25:
        font_size = 5.2
        header_font_size = 5.2
        top_bottom_padding = 2
        left_right_padding = 1
    elif num_cols > 15:
        font_size = 6
        header_font_size = 6.5
        top_bottom_padding = 3.5
        left_right_padding = 2
    elif num_cols > 8:
        font_size = 7.5
        header_font_size = 8
        top_bottom_padding = 5
        left_right_padding = 3
    else:
        font_size = 9
        header_font_size = 9.5
        top_bottom_padding = 6
        left_right_padding = 4
        
    # Scale down sizes if there are many rows to fit on page 1
    if num_rows > 45:
        font_size = min(font_size, 6.0)
        header_font_size = min(header_font_size, 6.5)
        top_bottom_padding = min(top_bottom_padding, 1.8)
    elif num_rows > 20:
        font_size = min(font_size, 7.5)
        header_font_size = min(header_font_size, 8.0)
        top_bottom_padding = min(top_bottom_padding, 2.5)

    # Professional Paragraph Styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=13,
        leading=15,
        textColor=colors.HexColor('#1e293b'),
        alignment=1, # Center
        spaceBefore=2,
        spaceAfter=4
    )
    
    cell_style = ParagraphStyle(
        'CustomCell',
        parent=styles['Normal'],
        fontSize=font_size,
        leading=font_size + 1.5,
        textColor=colors.HexColor('#334155'),
        alignment=1, # Center
        wordWrap='CJK'
    )
    
    header_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontSize=header_font_size,
        leading=header_font_size + 1.5,
        textColor=colors.white,
        alignment=1, # Center
        wordWrap='CJK'
    )
    
    # 1. Custom Header Layout (Logo/Tagline on Left, Company Details in Middle, Software Name on Right)
    logo = None
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Radiogeet blue logo.png')
    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path)
            aspect = logo.imageWidth / float(logo.imageHeight)
            logo.drawHeight = 0.5 * inch
            logo.drawWidth = (0.5 * inch) * aspect
            logo.hAlign = 'CENTER'
        except Exception:
            pass
            
    tagline_style = ParagraphStyle(
        'Tagline',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.black, # Darker color for bolder look
        alignment=1, # Center
        spaceBefore=0
    )
    
    company_name_style = ParagraphStyle(
        'CompanyName',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#991b1b'),
        alignment=1, # Center
        spaceAfter=0
    )
    company_address_style = ParagraphStyle(
        'CompanyAddress',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.HexColor('#475569'),
        alignment=1, # Center
        spaceAfter=2
    )
    company_sub_style = ParagraphStyle(
        'CompanySub',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.HexColor('#475569'),
        alignment=1, # Center
    )
    
    soft_title_style = ParagraphStyle(
        'SoftTitle',
        parent=styles['Heading2'],
        fontSize=15,
        textColor=colors.HexColor('#2563eb'), # Vibrant blue
        alignment=2, # Right aligned for equal edge
        spaceAfter=0
    )
    full_name_style = ParagraphStyle(
        'FullName',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.HexColor('#475569'),
        alignment=2, # Right aligned for equal edge
    )
    
    left_row0 = logo if logo else Paragraph("", tagline_style)
    middle_row0 = Paragraph("<b>Gemini Distilleries (Jharkhand) Pvt. Ltd.</b>", company_name_style)
    right_row0 = Paragraph("<b>Radiogeet AXIONIX</b>", soft_title_style)
    
    left_row1 = Paragraph("<b>Built in India | Innovated in Singapore | Delivered Globally</b>", tagline_style)
    middle_row1_elements = [
        Paragraph("Plot No. - 4 & 5, Tatisilwai Industrial Area,<br/>Phase - II, Tatisilwai, Ranchi,<br/>Pin - 835103 (Jharkhand)", company_address_style),
        Paragraph("<b>CIN No. :</b> U15549KA2000PTC028205", company_sub_style),
        Paragraph("<b>Email :</b> geminidistilleriesjharkhand@gmail.com", company_sub_style)
    ]
    right_row1 = Paragraph("Advanced X-Connected Intelligent Operational Network for Industrial eXcellence", full_name_style)
    
    header_data = [
        [left_row0, middle_row0, right_row0],
        [left_row1, middle_row1_elements, right_row1]
    ]
    
    header_table = Table(header_data, colWidths=[166, 230, 166])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'), # Center align Row 0 elements (logo and titles)
        ('VALIGN', (0, 1), (-1, 1), 'TOP'),    # Top align Row 1 elements (subtexts)
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Center align Left column (logo and tagline)
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Center align Middle column (company details)
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),   # Right align Right column (software title & description)
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 4))
            
    # 2. Add Document Title & Metadata
    elements.append(Paragraph(title, title_style))
    if metadata:
        meta_left_style = ParagraphStyle(
            'MetaLeft',
            parent=styles['Normal'],
            fontSize=7.5,
            textColor=colors.HexColor('#64748b'),
            alignment=0, # Left
        )
        meta_right_style = ParagraphStyle(
            'MetaRight',
            parent=styles['Normal'],
            fontSize=7.5,
            textColor=colors.HexColor('#64748b'),
            alignment=2, # Right
        )
        meta_left_p = Paragraph(f"<b>Generated On:</b> {metadata['generated_on']}", meta_left_style)
        meta_right_p = Paragraph(f"<b>Report Period:</b> {metadata['period']}", meta_right_style)
        
        meta_table = Table([[meta_left_p, meta_right_p]], colWidths=[281.0, 281.0])
        meta_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 5))
    
    # 3. Process Data for Text Wrapping & Build Section Tables
    printable_width = 562.0
    if num_cols > 1:
        date_col_width = 75.0 if num_cols > 20 else (90.0 if num_cols > 10 else 105.0)
        remaining_width = printable_width - date_col_width
        col_widths = [date_col_width] + [remaining_width / (num_cols - 1)] * (num_cols - 1)
    else:
        col_widths = [printable_width]

    sec_title_style = ParagraphStyle(
        'SecTitle',
        parent=styles['Heading2'],
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#1e293b'),
        alignment=0, # Left
        spaceBefore=6,
        spaceAfter=3
    )

    def append_pdf_table(table_heading, section_rows):
        if not section_rows:
            return
        if table_heading:
            elements.append(Paragraph(f"<b>{table_heading}</b>", sec_title_style))
            
        wrapped_headers = [Paragraph(f"<b>{h}</b>", header_style) for h in headers]
        wrapped_rows = []
        totals_indices = []
        for idx, row in enumerate(section_rows):
            is_tot = (len(row) > 0 and str(row[0]) in ('TANK TOTAL', 'FLOW METER TOTAL'))
            if is_tot:
                totals_indices.append(idx + 1)
            wrapped_row = []
            for cell in row:
                t = str(cell)
                if is_tot:
                    t = f"<b>{t}</b>"
                wrapped_row.append(Paragraph(t, cell_style))
            wrapped_rows.append(wrapped_row)
            
        sec_data = [wrapped_headers] + wrapped_rows
        
        sec_t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), top_bottom_padding + 2),
            ('TOPPADDING', (0, 0), (-1, 0), top_bottom_padding + 2),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('BOTTOMPADDING', (0, 1), (-1, -1), top_bottom_padding),
            ('TOPPADDING', (0, 1), (-1, -1), top_bottom_padding),
            ('LEFTPADDING', (0, 0), (-1, -1), left_right_padding),
            ('RIGHTPADDING', (0, 0), (-1, -1), left_right_padding),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]
        
        for r_idx in totals_indices:
            sec_t_style.append(('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#f1f5f9')))
            sec_t_style.append(('LINEABOVE', (0, r_idx), (-1, r_idx), 1, colors.HexColor('#cbd5e1')))
            sec_t_style.append(('LINEBELOW', (0, r_idx), (-1, r_idx), 1, colors.HexColor('#cbd5e1')))
            
        t_obj = Table(sec_data, colWidths=col_widths, repeatRows=1)
        t_obj.setStyle(TableStyle(sec_t_style))
        elements.append(t_obj)
        elements.append(Spacer(1, 6))

    if tank_rows or fm_rows:
        if tank_rows:
            append_pdf_table("Tank Details", tank_rows)
        if fm_rows:
            append_pdf_table("Flow Meter Details", fm_rows)
    else:
        # Fallback: Partition rows if TANK TOTAL or FLOW METER TOTAL are present
        t_rows = []
        f_rows = []
        found_tank_total = False
        for r in rows:
            if not found_tank_total:
                t_rows.append(r)
                if len(r) > 0 and r[0] == 'TANK TOTAL':
                    found_tank_total = True
            else:
                f_rows.append(r)
                
        if found_tank_total and f_rows:
            append_pdf_table("Tank Details", t_rows)
            append_pdf_table("Flow Meter Details", f_rows)
        else:
            append_pdf_table(None, rows)
    
    if change_headers and change_rows:
        elements.append(Spacer(1, 20))
        
        # Add Section Title
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#1e293b'),
            alignment=0, # Left
            spaceAfter=10
        )
        elements.append(Paragraph("<b>Device Data Change Log</b>", section_style))
        
        # Build changes table
        wrapped_change_headers = [Paragraph(f"<b>{h}</b>", header_style) for h in change_headers]
        wrapped_change_rows = []
        for row in change_rows:
            wrapped_row = [Paragraph(str(cell), cell_style) for cell in row]
            wrapped_change_rows.append(wrapped_row)
            
        c_data = [wrapped_change_headers] + wrapped_change_rows
        c_widths = [140.0, 90.0, 100.0, 400.0]
        
        c_table = Table(c_data, colWidths=c_widths, repeatRows=1)
        c_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#475569')), # Slate header
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), top_bottom_padding + 2),
            ('TOPPADDING', (0, 0), (-1, 0), top_bottom_padding + 2),
            
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ('BOTTOMPADDING', (0, 1), (-1, -1), top_bottom_padding),
            ('TOPPADDING', (0, 1), (-1, -1), top_bottom_padding),
            ('LEFTPADDING', (0, 0), (-1, -1), left_right_padding),
            ('RIGHTPADDING', (0, 0), (-1, -1), left_right_padding),
            
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ]))
        elements.append(c_table)
    
    def draw_page_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(1.5)
        width, height = doc.pagesize
        # Draw border outside the content margins (margins are 30, so border at 15)
        margin = 15
        canvas.rect(margin, margin, width - (margin * 2), height - (margin * 2))
        
        # Add copyright footer
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748b'))
        canvas.drawCentredString(width / 2.0, margin + 5, "© Copyright: Radiogeet Digital Pvt. Ltd. | Radiogeet AXIONIX")
        
        canvas.restoreState()
        
    doc.build(elements, onFirstPage=draw_page_border, onLaterPages=draw_page_border)
    
    return response


def generate_excel_response(headers, rows, title, filename, metadata=None, change_headers=None, change_rows=None, tank_rows=None, fm_rows=None):
    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    from django.conf import settings
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Data"

    # Logo and header info
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Radiogeet blue logo.png')
    
    # Try inserting logo
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            original_width = img.width
            original_height = img.height
            img.height = 70
            img.width = int(original_width * (70 / original_height))
            ws.add_image(img, 'A1')
        except Exception:
            pass

    from openpyxl.utils import get_column_letter
    max_col_letter = get_column_letter(max(8, len(headers)))

    # Set row heights for elegant spacing
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # Company Details in Middle (columns C to E, rows 1 to 3)
    ws.merge_cells('C1:E1')
    cell_comp_name = ws['C1']
    cell_comp_name.value = "Gemini Distilleries (Jharkhand) Pvt. Ltd."
    cell_comp_name.font = Font(size=12, bold=True, color="991B1B")
    cell_comp_name.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('C2:E2')
    cell_comp_addr = ws['C2']
    cell_comp_addr.value = "Plot No. - 4 & 5, Tatisilwai Industrial Area, Phase - II, Tatisilwai, Ranchi, Pin - 835103 (Jharkhand)"
    cell_comp_addr.font = Font(size=8, color="475569")
    cell_comp_addr.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('C3:E3')
    cell_comp_contact = ws['C3']
    cell_comp_contact.value = "CIN No. : U15549KA2000PTC028205  |  Email : geminidistilleriesjharkhand@gmail.com"
    cell_comp_contact.font = Font(size=8, color="475569")
    cell_comp_contact.alignment = Alignment(horizontal="center", vertical="center")

    # Software Name and Tagline on the right (columns F to max_col_letter)
    ws.merge_cells(f'F1:{max_col_letter}1')
    cell_soft = ws['F1']
    cell_soft.value = "Radiogeet AXIONIX"
    cell_soft.font = Font(size=15, bold=True, color="2563EB")
    cell_soft.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f'F2:{max_col_letter}2')
    cell_tag = ws['F2']
    cell_tag.value = "Advanced X-Connected Intelligent Operational Network for Industrial eXcellence"
    cell_tag.font = Font(size=9, color="475569")
    cell_tag.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f'F3:{max_col_letter}3')
    cell_tagline = ws['F3']
    cell_tagline.value = "Built in India | Innovated in Singapore | Delivered Globally"
    cell_tagline.font = Font(size=8, color="475569")
    cell_tagline.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f'A4:{max_col_letter}4')
    cell_title = ws['A4']
    cell_title.value = title
    cell_title.font = Font(size=14, bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    if metadata:
        half_col_num = max(8, len(headers)) // 2
        half_letter = get_column_letter(half_col_num)
        start_right_letter = get_column_letter(half_col_num + 1)
        
        ws.merge_cells(f'A5:{half_letter}5')
        cell_meta_left = ws['A5']
        cell_meta_left.value = f"Generated On: {metadata['generated_on']}"
        cell_meta_left.font = Font(size=9, italic=True, color="64748B")
        cell_meta_left.alignment = Alignment(horizontal="left", vertical="center")
        
        ws.merge_cells(f'{start_right_letter}5:{max_col_letter}5')
        cell_meta_right = ws[f'{start_right_letter}5']
        cell_meta_right.value = f"Report Period: {metadata['period']}"
        cell_meta_right.font = Font(size=9, italic=True, color="64748B")
        cell_meta_right.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.row_dimensions[5].height = 18
        start_row = 7
    else:
        start_row = 6

    # Headers
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows
    last_row = start_row
    for row_num, row_data in enumerate(rows, start_row + 1):
        last_row = row_num
        is_totals_row = (len(row_data) > 0 and row_data[0] in ('TANK TOTAL', 'FLOW METER TOTAL'))
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
            
            # Formatting
            if is_totals_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            else:
                cell.font = Font(bold=False)
                
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Add change log if provided
    if change_headers and change_rows:
        start_change_row = last_row + 4
        ws.merge_cells(f'A{start_change_row}:{max_col_letter}{start_change_row}')
        cell_sec_title = ws[f'A{start_change_row}']
        cell_sec_title.value = "Device Data Change Log"
        cell_sec_title.font = Font(size=12, bold=True, color="475569")
        cell_sec_title.alignment = Alignment(horizontal="left", vertical="center")
        
        header_row = start_change_row + 1
        change_header_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        for col_num, header in enumerate(change_headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = change_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        last_change_row = header_row
        for row_num, row_data in enumerate(change_rows, header_row + 1):
            last_change_row = row_num
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        copy_row = last_change_row + 2
    else:
        copy_row = last_row + 2

    # Add copyright at the bottom
    ws.merge_cells(f'A{copy_row}:{max_col_letter}{copy_row}')
    cell_copy = ws[f'A{copy_row}']
    cell_copy.value = "© Copyright: Radiogeet Digital Pvt. Ltd. | Built in India | Delivered Globally"
    cell_copy.font = Font(size=9, bold=True, color="1E293B")
    cell_copy.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_num in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.row >= start_row and cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 40:
            adjusted_width = 40
        elif adjusted_width < 12:
            adjusted_width = 12
        ws.column_dimensions[column_letter].width = adjusted_width

    from io import BytesIO
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)

    response = HttpResponse(virtual_workbook.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_csv_response(headers, rows, title, filename, metadata=None, change_headers=None, change_rows=None, tank_rows=None, fm_rows=None):
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Branded text logo and metadata header matching Excel style
    writer.writerow(["=================================================="])
    writer.writerow(["RADIOGEET AXIONIX - REPORT SYSTEM"])
    writer.writerow(["Tagline: Advanced X-Connected Intelligent Operational Network for Industrial eXcellence"])
    writer.writerow(["--------------------------------------------------"])
    writer.writerow(["Gemini Distilleries (Jharkhand) Pvt. Ltd."])
    writer.writerow(["Plot No. - 4 & 5, Tatisilwai Industrial Area, Phase - II, Tatisilwai, Ranchi, Pin - 835103 (Jharkhand)"])
    writer.writerow(["CIN No. : U15549KA2000PTC028205 | Email : geminidistilleriesjharkhand@gmail.com"])
    writer.writerow(["=================================================="])
    writer.writerow([f"Report Name: {title}"])
    if metadata:
        writer.writerow([f"Report Period: {metadata['period']}"])
        writer.writerow([f"Generated On: {metadata['generated_on']}"])
    else:
        writer.writerow([f"Generated At: {timezone.now().strftime('%Y-%m-%d %I:%M:%S %p')}"])
    writer.writerow([])
    
    if tank_rows or fm_rows:
        if tank_rows:
            writer.writerow(["--- TANK DETAILS ---"])
            writer.writerow(headers)
            for r in tank_rows:
                writer.writerow(r)
            writer.writerow([])
        if fm_rows:
            writer.writerow(["--- FLOW METER DETAILS ---"])
            writer.writerow(headers)
            for r in fm_rows:
                writer.writerow(r)
            writer.writerow([])
    else:
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

    if change_headers and change_rows:
        writer.writerow([])
        writer.writerow(["=================================================="])
        writer.writerow(["DEVICE DATA CHANGE LOG"])
        writer.writerow(["=================================================="])
        writer.writerow([])
        writer.writerow(change_headers)
        for row in change_rows:
            writer.writerow(row)
        
    writer.writerow([])
    writer.writerow(["Software Name:", "Radiogeet AXIONIX"])
    writer.writerow(["Copyright:", "© Copyright: Radiogeet Digital Pvt. Ltd. | Built in India | Delivered Globally"])
    
    return response



@login_required
def export_trend_csv(request):
    """Exports historical tank and flow meter trend data matrix as CSV, Excel, or PDF."""
    from django.utils import timezone
    from datetime import timedelta
    from django.utils.dateparse import parse_datetime
    
    fmt = request.GET.get('format', 'csv')
    range_param = request.GET.get('range', '6H').upper()
    device_id = request.GET.get('device_id', 'all')
    
    all_tanks = get_user_tanks(request.user).order_by("tank_id")
    all_fms = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    if device_id.startswith('tank_'):
        t_id = device_id.split('_', 1)[1]
        tanks = list(all_tanks.filter(tank_id__iexact=t_id))
        flow_meters = []
    elif device_id.startswith('fm_'):
        fm_id = device_id.split('_', 1)[1]
        tanks = []
        flow_meters = list(all_fms.filter(meter_id__iexact=fm_id))
    elif all_tanks.filter(tank_id__iexact=device_id).exists():
        tanks = list(all_tanks.filter(tank_id__iexact=device_id))
        flow_meters = []
    elif all_fms.filter(meter_id__iexact=device_id).exists():
        tanks = []
        flow_meters = list(all_fms.filter(meter_id__iexact=device_id))
    else:
        tanks = list(all_tanks)
        flow_meters = list(all_fms)

    now = timezone.now()
    if range_param == "1H":
        start_time = now - timedelta(hours=1)
    elif range_param == "12H":
        start_time = now - timedelta(hours=12)
    elif range_param == "24H":
        start_time = now - timedelta(hours=24)
    elif range_param == "7D":
        start_time = now - timedelta(days=7)
    elif range_param == "CUSTOM":
        custom_start = request.GET.get("start_time")
        custom_end = request.GET.get("end_time")
        start_time = None
        if custom_start:
            try:
                from datetime import datetime
                naive_start = datetime.strptime(custom_start, "%Y-%m-%dT%H:%M")
                start_time = timezone.make_aware(naive_start, timezone.get_current_timezone())
            except ValueError:
                dt_start = parse_datetime(custom_start)
                if dt_start and timezone.is_naive(dt_start):
                    start_time = timezone.make_aware(dt_start, timezone.get_current_timezone())
                elif dt_start:
                    start_time = dt_start

        if custom_end:
            try:
                from datetime import datetime
                naive_end = datetime.strptime(custom_end, "%Y-%m-%dT%H:%M")
                now = timezone.make_aware(naive_end, timezone.get_current_timezone())
            except ValueError:
                dt_end = parse_datetime(custom_end)
                if dt_end and timezone.is_naive(dt_end):
                    now = timezone.make_aware(dt_end, timezone.get_current_timezone())
                elif dt_end:
                    now = dt_end

        if not start_time:
            start_time = now - timedelta(hours=6)
    else: # 6H
        start_time = now - timedelta(hours=6)

    tank_readings = list(TankReading.objects.filter(tank__in=tanks, timestamp__gte=start_time, timestamp__lte=now).select_related("tank").order_by("timestamp"))
    fm_readings = list(FlowMeterReading.objects.filter(flow_meter__in=flow_meters, timestamp__gte=start_time, timestamp__lte=now).select_related("flow_meter").order_by("timestamp"))

    headers = ['Date & Time']
    for t in tanks:
        unit_label = t.flow_unit if t.widget_type == 'flow_meter' else t.unit
        if t.name == t.tank_id:
            headers.append(f"{t.tank_id} ({unit_label})")
        else:
            headers.append(f"{t.name} ({unit_label})")
    for fm in flow_meters:
        if fm.name == fm.meter_id:
            headers.append(f"{fm.meter_id} Rate")
        else:
            headers.append(f"{fm.name} Rate")

    time_groups = {}
    for r in tank_readings:
        local_dt = timezone.localtime(r.timestamp)
        bucket_dt = local_dt.replace(microsecond=0)
        t_str = bucket_dt.strftime("%d %b %Y %I:%M:%S %p")
        if t_str not in time_groups:
            time_groups[t_str] = {}
        if r.tank.widget_type == 'flow_meter':
            time_groups[t_str][r.tank.tank_id] = f"{r.flow_rate} {r.tank.flow_unit}" if r.flow_rate is not None else "--"
        else:
            time_groups[t_str][r.tank.tank_id] = format_tank_level(r.level_percent, r.tank.capacity_kl, r.tank.unit, getattr(r, 'raw_value', None))

    for r in fm_readings:
        local_dt = timezone.localtime(r.timestamp)
        bucket_dt = local_dt.replace(microsecond=0)
        t_str = bucket_dt.strftime("%d %b %Y %I:%M:%S %p")
        if t_str not in time_groups:
            time_groups[t_str] = {}
        time_groups[t_str][f"fm_{r.flow_meter.meter_id}"] = f"{r.flow_rate} {r.flow_meter.flow_unit}"

    sorted_times = sorted(time_groups.keys(), reverse=True)
    rows = []
    for t_str in sorted_times:
        row = [t_str]
        for t in tanks:
            val = time_groups[t_str].get(t.tank_id, "--")
            row.append(val)
        for fm in flow_meters:
            val = time_groups[t_str].get(f"fm_{fm.meter_id}", "--")
            row.append(val)
        rows.append(row)

    dev_label = device_id if device_id != 'all' else 'All Devices'
    report_title = f"Tank & Flow Trend Data Matrix (Range: {range_param}, Device: {dev_label})"

    if fmt == 'pdf':
        return generate_pdf_response(headers, rows, report_title, "tank_levels_trend_export.pdf")
    elif fmt == 'excel':
        return generate_excel_response(headers, rows, report_title, "tank_levels_trend_export.xlsx")
        
    return generate_csv_response(headers, rows, report_title, "tank_levels_trend_export.csv")


@login_required
def export_alarms_csv(request):
    """Exports active and historical alarms list as CSV, Excel, or PDF."""
    from django.db.models import Q
    
    fmt = request.GET.get('format', 'csv')
    alarm_type = request.GET.get('alarm_type', 'all').strip()
    status = request.GET.get('status', 'all').strip()
    q = request.GET.get('q', '').strip()
    
    headers = ['Timestamp', 'Device ID', 'Device Name', 'Alarm Type', 'Severity', 'Value', 'Status', 'Message']
    rows = []
    
    alarms_qs = Alarm.objects.select_related('tank', 'flow_meter').all()
    if alarm_type and alarm_type.lower() != 'all':
        if 'high' in alarm_type.lower():
            alarms_qs = alarms_qs.filter(alarm_type='high')
        elif 'low' in alarm_type.lower():
            alarms_qs = alarms_qs.filter(alarm_type='low')

    if status and status.lower() != 'all':
        if status.lower() in ['active', 'unacknowledged']:
            alarms_qs = alarms_qs.filter(acknowledged=False)
        elif status.lower() == 'acknowledged':
            alarms_qs = alarms_qs.filter(acknowledged=True)

    if q:
        alarms_qs = alarms_qs.filter(
            Q(tank__tank_id__icontains=q) |
            Q(tank__name__icontains=q) |
            Q(flow_meter__meter_id__icontains=q) |
            Q(flow_meter__name__icontains=q) |
            Q(message__icontains=q)
        )

    alarms = alarms_qs.order_by('-timestamp')[:1000]
    
    for a in alarms:
        status_str = "Acknowledged" if a.acknowledged else "Active"
        device_id = a.tank.tank_id if a.tank else (a.flow_meter.meter_id if a.flow_meter else "Unknown")
        device_name = a.tank.name if a.tank else (a.flow_meter.name if a.flow_meter else "Unknown")
        
        if a.tank:
            if a.tank.unit == "RAW":
                val_str = format_tank_level(None, a.tank.capacity_kl, a.tank.unit, a.level_percent)
            else:
                val_str = format_tank_level(a.level_percent, a.tank.capacity_kl, a.tank.unit)
            high_limit = format_tank_level(a.tank.high_limit, a.tank.capacity_kl, a.tank.unit, a.tank.high_limit)
            low_limit = format_tank_level(a.tank.low_limit, a.tank.capacity_kl, a.tank.unit, a.tank.low_limit)
            msg_str = f"Value exceeded High Limit ({high_limit})" if a.alarm_type == "high" else f"Value dropped below Low Limit ({low_limit})"
        elif a.flow_meter:
            val_str = f"{a.level_percent:.1f} {a.flow_meter.flow_unit}"
            high_limit = f"{a.flow_meter.high_limit} {a.flow_meter.flow_unit}"
            low_limit = f"{a.flow_meter.low_limit} {a.flow_meter.flow_unit}"
            msg_str = f"Flow rate exceeded High Limit ({high_limit})" if a.alarm_type == "high" else f"Flow rate dropped below Low Limit ({low_limit})"
        else:
            val_str = f"{a.level_percent:.1f}"
            msg_str = getattr(a, 'message', f"Value exceeded limit ({val_str})")
            
        severity_str = getattr(a, 'severity', 'HIGH' if a.alarm_type == 'high' else 'WARNING').upper()
        rows.append([
            a.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            device_id,
            device_name,
            a.get_alarm_type_display(),
            severity_str,
            val_str,
            status_str,
            msg_str
        ])
        
    filter_desc = []
    if alarm_type and alarm_type.lower() != 'all':
        filter_desc.append(f"Type: {alarm_type}")
    if status and status.lower() != 'all':
        filter_desc.append(f"Status: {status.title()}")
    if q:
        filter_desc.append(f"Search: '{q}'")
        
    title_suffix = f" ({', '.join(filter_desc)})" if filter_desc else " (All Alarms)"
    report_title = f"Alarms History Log{title_suffix}"

    if fmt == 'pdf':
        return generate_pdf_response(headers, rows, report_title, "alarms_history_export.pdf")
    elif fmt == 'excel':
        return generate_excel_response(headers, rows, report_title, "alarms_history_export.xlsx")
        
    return generate_csv_response(headers, rows, report_title, "alarms_history_export.csv")


@login_required
def export_events_csv(request):
    """Exports system operational events log as CSV, Excel, or PDF."""
    from .models import EventLog
    from django.db.models import Q
    
    fmt = request.GET.get('format', 'csv')
    event_type = request.GET.get('event_type', 'all').strip()
    source = request.GET.get('source', 'all').strip()
    q = request.GET.get('q', '').strip()

    headers = ['Timestamp', 'Event Type', 'Source', 'Log Message']
    rows = []
    
    logs_qs = EventLog.objects.all()
    if event_type and event_type.lower() != 'all':
        logs_qs = logs_qs.filter(event_type__iexact=event_type)
    if source and source.lower() != 'all':
        logs_qs = logs_qs.filter(source__icontains=source)
    if q:
        logs_qs = logs_qs.filter(Q(message__icontains=q) | Q(source__icontains=q))

    logs = logs_qs.order_by('-timestamp')[:1000]
    for log in logs:
        rows.append([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.event_type,
            log.source,
            log.message
        ])
        
    filter_desc = []
    if event_type and event_type.lower() != 'all':
        filter_desc.append(f"Type: {event_type.upper()}")
    if source and source.lower() != 'all':
        filter_desc.append(f"Source: {source}")
    if q:
        filter_desc.append(f"Search: '{q}'")
        
    title_suffix = f" ({', '.join(filter_desc)})" if filter_desc else " (All Events)"
    report_title = f"System Events Log{title_suffix}"

    if fmt == 'pdf':
        return generate_pdf_response(headers, rows, report_title, "system_events_log.pdf")
    elif fmt == 'excel':
        return generate_excel_response(headers, rows, report_title, "system_events_log.xlsx")
        
    return generate_csv_response(headers, rows, report_title, "system_events_log.csv")


@login_required
def export_reports_csv(request):
    """Exports aggregated tank level reports summary as CSV."""
    import csv
    from django.http import HttpResponse
    from django.db.models import Avg, Min, Max
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Determine active tanks based on user profile assignment
    all_tanks_list = get_user_tanks(request.user).order_by("tank_id")
    all_flow_meters_list = FlowMeter.objects.filter(is_active=True).order_by("meter_id")

    # Parse GET filters
    report_type = request.GET.get("report_type", "daily").lower()
    if report_type not in ["daily", "weekly", "monthly", "custom"]:
        report_type = "daily"

    device_id = request.GET.get("device_id", "all")
    if device_id.startswith("tank_"):
        t_id = device_id.split("_")[1]
        tanks = all_tanks_list.filter(tank_id=t_id)
        flow_meters = FlowMeter.objects.none()
    elif device_id.startswith("fm_"):
        fm_id = device_id.split("_")[1]
        tanks = Tank.objects.none()
        flow_meters = all_flow_meters_list.filter(meter_id=fm_id)
    else:
        tanks = all_tanks_list
        flow_meters = all_flow_meters_list

    # Parse base date
    date_str = request.GET.get("date", "")
    try:
        base_date = datetime.strptime(date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        base_date = timezone.now()
    
    # Parse end date for custom reports
    end_date_str = request.GET.get("end_date", "")
    try:
        custom_end = datetime.strptime(end_date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        custom_end = base_date

    # Determine start and end ranges
    if report_type == "weekly":
        start_dt = base_date - timedelta(days=6)
        end_dt = base_date
    elif report_type == "monthly":
        start_dt = base_date - timedelta(days=29)
        end_dt = base_date
    elif report_type == "custom":
        start_dt = base_date
        end_dt = custom_end
    else:  # daily
        start_dt = base_date
        end_dt = base_date

    # Normalize to start and end of day
    start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Convert to timezone aware if required
    from django.conf import settings
    if settings.USE_TZ:
        if timezone.is_naive(start_dt):
            start_dt = timezone.make_aware(start_dt)
        if timezone.is_naive(end_dt):
            end_dt = timezone.make_aware(end_dt)

    fmt = request.GET.get('format', 'csv')
    
    headers = [
        'Device ID', 'Device Name', 'Opening Value', 'Closing Value', 'Difference'
    ]
    tank_rows = []
    fm_rows = []
    
    # 1. Regular Tanks
    tank_open_sum = 0
    tank_close_sum = 0
    regular_tanks_count = 0
    primary_tank_unit = "RAW"
    
    regular_tanks = [t for t in tanks if t.widget_type != 'flow_meter']
    for tank in regular_tanks:
        primary_tank_unit = tank.unit or "RAW"
        readings = TankReading.objects.filter(tank=tank, timestamp__range=(start_dt, end_dt))
        
        if readings.exists():
            oldest = readings.order_by("timestamp").first()
            newest = readings.order_by("-timestamp").first()
            opening_val = format_tank_level(oldest.level_percent, tank.capacity_kl, tank.unit, oldest.raw_value) if oldest else "--"
            closing_val = format_tank_level(newest.level_percent, tank.capacity_kl, tank.unit, newest.raw_value) if newest else "--"
            
            diff_pct = newest.level_percent - oldest.level_percent if (newest and oldest) else 0.0
            diff_raw = newest.raw_value - oldest.raw_value if (newest and oldest and newest.raw_value is not None and oldest.raw_value is not None) else None
            diff_val_str = format_tank_level(diff_pct, tank.capacity_kl, tank.unit, diff_raw)
            
            if oldest and newest:
                if tank.unit == "RAW":
                    open_v = oldest.raw_value if oldest.raw_value is not None else 0.0
                    close_v = newest.raw_value if newest.raw_value is not None else 0.0
                else:
                    open_v = get_tank_value_in_unit(oldest.level_percent, tank.capacity_kl, tank.unit, oldest.raw_value)
                    close_v = get_tank_value_in_unit(newest.level_percent, tank.capacity_kl, tank.unit, newest.raw_value)
                tank_open_sum += open_v
                tank_close_sum += close_v
                regular_tanks_count += 1
        else:
            opening_val = "--"
            closing_val = "--"
            diff_val_str = "--"
            
        tank_rows.append([
            tank.tank_id,
            tank.name,
            opening_val,
            closing_val,
            diff_val_str
        ])
        
    if len(tank_rows) > 0:
        tank_diff_sum = tank_close_sum - tank_open_sum
        if regular_tanks_count > 0:
            tank_open_str = format_total_val(tank_open_sum, primary_tank_unit)
            tank_close_str = format_total_val(tank_close_sum, primary_tank_unit)
            tank_diff_str = format_total_val(tank_diff_sum, primary_tank_unit)
            if primary_tank_unit != "RAW" and tank_diff_sum >= 0 and not tank_diff_str.startswith("+"):
                tank_diff_str = f"+{tank_diff_str}"
        else:
            tank_open_str = "--"
            tank_close_str = "--"
            tank_diff_str = "--"
            
        tank_rows.append([
            'TANK TOTAL',
            '-',
            tank_open_str,
            tank_close_str,
            tank_diff_str
        ])
        
    # 2. Flow Meters (Flow Meter Tanks + Standalone FlowMeters)
    fm_open_sum = 0
    fm_close_sum = 0
    fm_count = 0
    primary_fm_unit = "Liters"
    
    fm_tanks = [t for t in tanks if t.widget_type == 'flow_meter']
    for tank in fm_tanks:
        primary_fm_unit = tank.total_unit or "Liters"
        readings = TankReading.objects.filter(tank=tank, timestamp__range=(start_dt, end_dt))
        
        if readings.exists():
            oldest = readings.order_by("timestamp").first()
            newest = readings.order_by("-timestamp").first()
            opening_val = oldest.total_flow if (oldest and oldest.total_flow is not None) else 0.0
            closing_val = newest.total_flow if (newest and newest.total_flow is not None) else 0.0
            diff_val = closing_val - opening_val
            
            opening_val_str = f"{round(opening_val, 2)} {tank.total_unit}"
            closing_val_str = f"{round(closing_val, 2)} {tank.total_unit}"
            diff_val_str = f"{round(diff_val, 2)} {tank.total_unit}"
            
            fm_open_sum += opening_val
            fm_close_sum += closing_val
            fm_count += 1
        else:
            opening_val_str = "--"
            closing_val_str = "--"
            diff_val_str = "--"
            
        fm_rows.append([
            tank.tank_id,
            tank.name,
            opening_val_str,
            closing_val_str,
            diff_val_str
        ])
        
    for fm in flow_meters:
        unit_str = fm.total_unit or fm.flow_unit or "m³/h"
        primary_fm_unit = unit_str
        readings = FlowMeterReading.objects.filter(flow_meter=fm, timestamp__range=(start_dt, end_dt))
        
        if readings.exists():
            oldest = readings.order_by("timestamp").first()
            newest = readings.order_by("-timestamp").first()
            
            opening_val = oldest.total_volume if oldest else 0.0
            closing_val = newest.total_volume if newest else 0.0
            diff_val = closing_val - opening_val
            
            opening_val_str = f"{round(opening_val, 2)} {unit_str}"
            closing_val_str = f"{round(closing_val, 2)} {unit_str}"
            diff_val_str = f"{round(diff_val, 2)} {unit_str}"
            
            fm_open_sum += opening_val
            fm_close_sum += closing_val
            fm_count += 1
        else:
            opening_val_str = "--"
            closing_val_str = "--"
            diff_val_str = "--"
            
        fm_rows.append([
            fm.meter_id,
            fm.name,
            opening_val_str,
            closing_val_str,
            diff_val_str
        ])
        
    if len(fm_rows) > 0:
        fm_diff_sum = fm_close_sum - fm_open_sum
        if fm_count > 0:
            fm_open_str = f"{round(fm_open_sum, 2)} {primary_fm_unit}"
            fm_close_str = f"{round(fm_close_sum, 2)} {primary_fm_unit}"
            fm_diff_str = f"{round(fm_diff_sum, 2)} {primary_fm_unit}"
            if fm_diff_sum >= 0 and not fm_diff_str.startswith("+"):
                fm_diff_str = f"+{fm_diff_str}"
        else:
            fm_open_str = "--"
            fm_close_str = "--"
            fm_diff_str = "--"
            
        fm_rows.append([
            'FLOW METER TOTAL',
            '-',
            fm_open_str,
            fm_close_str,
            fm_diff_str
        ])
        
    metadata = {
        'generated_on': timezone.localtime(timezone.now()).strftime('%d %b %Y %I:%M:%S %p'),
        'period': f"{timezone.localtime(start_dt).strftime('%d %b %Y')} to {timezone.localtime(end_dt).strftime('%d %b %Y')}"
    }
    
    report_title = f"Device {report_type.capitalize()} Reports Summary"
        
    all_rows = tank_rows + fm_rows
    if fmt == 'pdf':
        return generate_pdf_response(headers, all_rows, report_title, f"device_{report_type}_reports.pdf", metadata=metadata, tank_rows=tank_rows, fm_rows=fm_rows)
    elif fmt == 'excel':
        return generate_excel_response(headers, all_rows, report_title, f"device_{report_type}_reports.xlsx", metadata=metadata, tank_rows=tank_rows, fm_rows=fm_rows)
        
    return generate_csv_response(headers, all_rows, report_title, f"device_{report_type}_reports.csv", metadata=metadata, tank_rows=tank_rows, fm_rows=fm_rows)


@login_required
def export_datalog_csv(request):
    """Exports the matrix log records for configured tanks as CSV."""
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    
    fmt = request.GET.get('format', 'csv')
    device_id = request.GET.get('device_id', 'all')
    try:
        interval_min = int(request.GET.get('interval', '1'))
    except (ValueError, TypeError):
        interval_min = 1
    
    all_tanks = get_user_tanks(request.user).order_by("tank_id")
    all_fms = FlowMeter.objects.filter(is_active=True).order_by('meter_id')
            
    if device_id.startswith('tank_'):
        t_id = device_id.split('_', 1)[1]
        tanks = list(all_tanks.filter(tank_id__iexact=t_id))
        flow_meters = []
    elif device_id.startswith('fm_'):
        fm_id = device_id.split('_', 1)[1]
        tanks = []
        flow_meters = list(all_fms.filter(meter_id__iexact=fm_id))
    elif all_tanks.filter(tank_id__iexact=device_id).exists():
        tanks = list(all_tanks.filter(tank_id__iexact=device_id))
        flow_meters = []
    elif all_fms.filter(meter_id__iexact=device_id).exists():
        tanks = []
        flow_meters = list(all_fms.filter(meter_id__iexact=device_id))
    else:
        tanks = list(all_tanks)
        flow_meters = list(all_fms)
    
    # Headers: Date & Time, Tank ID 1, Tank ID 2, ...
    headers = ['Date & Time']
    for tank in tanks:
        if tank.widget_type == 'flow_meter':
            if tank.name == tank.tank_id:
                headers.append(f"{tank.tank_id} Rate")
                headers.append(f"{tank.tank_id} Total")
            else:
                headers.append(f"{tank.name} Rate")
                headers.append(f"{tank.name} Total")
        else:
            if tank.name == tank.tank_id:
                headers.append(f"{tank.tank_id} ({tank.unit})")
            else:
                headers.append(f"{tank.name} ({tank.unit})")
    for fm in flow_meters:
        if fm.name == fm.meter_id:
            headers.append(f"{fm.meter_id} Rate")
            headers.append(f"{fm.meter_id} Total")
        else:
            headers.append(f"{fm.name} Rate")
            headers.append(f"{fm.name} Total")
        
    rows = []
    
    # Query last 1000 level history logs from database
    db_readings = list(TankReading.objects.filter(tank__in=tanks).select_related("tank").order_by("-timestamp")[:1000])
    db_fm_readings = list(FlowMeterReading.objects.filter(flow_meter__in=flow_meters).select_related("flow_meter").order_by("-timestamp")[:1000])
    
    # Compile readings into time logs matrix rows by interval resolution
    time_groups = {}
    for r in db_readings:
        local_dt = timezone.localtime(r.timestamp)
        if interval_min > 1:
            minute = (local_dt.minute // interval_min) * interval_min
            bucket_dt = local_dt.replace(minute=minute, second=0, microsecond=0)
            time_str = bucket_dt.strftime("%d %b %Y %I:%M %p")
        else:
            bucket_dt = local_dt.replace(microsecond=0)
            time_str = bucket_dt.strftime("%d %b %Y %I:%M:%S %p")
            
        if time_str not in time_groups:
            time_groups[time_str] = {}
        if r.tank.widget_type == 'flow_meter':
            if f"tank_{r.tank.tank_id}_rate" not in time_groups[time_str]:
                time_groups[time_str][f"tank_{r.tank.tank_id}_rate"] = f"{r.flow_rate} {r.tank.flow_unit}" if r.flow_rate is not None else "--"
                time_groups[time_str][f"tank_{r.tank.tank_id}_total"] = f"{r.total_flow} {r.tank.total_unit}" if r.total_flow is not None else "--"
        else:
            if r.tank.tank_id not in time_groups[time_str]:
                time_groups[time_str][r.tank.tank_id] = format_tank_level(r.level_percent, r.tank.capacity_kl, r.tank.unit, getattr(r, 'raw_value', None))
        
    for r in db_fm_readings:
        local_dt = timezone.localtime(r.timestamp)
        if interval_min > 1:
            minute = (local_dt.minute // interval_min) * interval_min
            bucket_dt = local_dt.replace(minute=minute, second=0, microsecond=0)
            time_str = bucket_dt.strftime("%d %b %Y %I:%M %p")
        else:
            bucket_dt = local_dt.replace(microsecond=0)
            time_str = bucket_dt.strftime("%d %b %Y %I:%M:%S %p")
            
        if time_str not in time_groups:
            time_groups[time_str] = {}
        if f"fm_{r.flow_meter.meter_id}_rate" not in time_groups[time_str]:
            time_groups[time_str][f"fm_{r.flow_meter.meter_id}_rate"] = f"{r.flow_rate} {r.flow_meter.flow_unit}"
            time_groups[time_str][f"fm_{r.flow_meter.meter_id}_total"] = f"{r.total_volume} {r.flow_meter.total_unit}"

    # Sort timestamps in reverse chronological order
    sorted_times = sorted(time_groups.keys(), reverse=True)
    
    for time_str in sorted_times:
        row = [time_str]
        for tank in tanks:
            if tank.widget_type == 'flow_meter':
                rate = time_groups[time_str].get(f"tank_{tank.tank_id}_rate", "--")
                total = time_groups[time_str].get(f"tank_{tank.tank_id}_total", "--")
                row.append(rate)
                row.append(total)
            else:
                val = time_groups[time_str].get(tank.tank_id, "--")
                row.append(val)
        for fm in flow_meters:
            rate = time_groups[time_str].get(f"fm_{fm.meter_id}_rate", "--")
            total = time_groups[time_str].get(f"fm_{fm.meter_id}_total", "--")
            row.append(rate)
            row.append(total)
        rows.append(row)
        
    dev_label = device_id if device_id != 'all' else 'All Devices'
    report_title = f"Data Log History Matrix (Device: {dev_label}, Interval: {interval_min} Mins)"
        
    if fmt == 'pdf':
        return generate_pdf_response(headers, rows, report_title, "data_log_history_matrix.pdf")
    elif fmt == 'excel':
        return generate_excel_response(headers, rows, report_title, "data_log_history_matrix.xlsx")
        
    return generate_csv_response(headers, rows, report_title, "data_log_history_matrix.csv")


@login_required
def export_events_csv(request):
    """Exports system operational events log as CSV."""
    import csv
    from django.http import HttpResponse
    from .models import EventLog
    
    fmt = request.GET.get('format', 'csv')
    headers = ['Timestamp', 'Event Type', 'Source', 'Log Message']
    rows = []
    
    logs = EventLog.objects.all().order_by('-timestamp')[:1000]
    for log in logs:
        rows.append([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.event_type,
            log.source,
            log.message
        ])
        
    if fmt == 'pdf':
        return generate_pdf_response(headers, rows, "System Events Log", "system_events_log.pdf")
    elif fmt == 'excel':
        return generate_excel_response(headers, rows, "System Events Log", "system_events_log.xlsx")
        
    return generate_csv_response(headers, rows, "System Events Log", "system_events_log.csv")


# =========================================
# User Management Views
# =========================================

MODULE_LIST = [
    {"key": "dashboard", "name": "Dashboard"},
    {"key": "trend_chart", "name": "Trend Chart"},
    {"key": "alarms", "name": "Alarms"},
    {"key": "reports", "name": "Reports"},
    {"key": "data_log", "name": "Data Log"},
    {"key": "events", "name": "Events"},
    {"key": "settings", "name": "Settings"},
    {"key": "users", "name": "Users"},
    {"key": "user_management", "name": "User Management"},
    {"key": "system_settings", "name": "System Settings"},
]


@login_required
@check_license_expiry
def users_list_view(request):
    """Page: User list."""
    users = UserProfile.objects.select_related("user").all().order_by("-user__date_joined")
    context = {
        "users": users,
        "current_time": timezone.now(),
        "user": request.user,
    }
    return render(request, "core/users_list.html", context)


def _build_module_context(profile=None):
    """Build the module permissions context list for the add/edit form."""
    modules = []
    for mod in MODULE_LIST:
        perm = None
        if profile:
            perm = ModulePermission.objects.filter(profile=profile, module_name=mod["key"]).first()
        modules.append({
            "key": mod["key"],
            "name": mod["name"],
            "can_view": perm.can_view if perm else True,
            "can_add": perm.can_add if perm else False,
            "can_edit": perm.can_edit if perm else False,
            "can_delete": perm.can_delete if perm else False,
            "can_export": perm.can_export if perm else False,
        })
    return modules


def _save_user_from_post(request, user_obj=None, profile_obj=None):
    """Create or update a user + profile + permissions from POST data."""
    full_name = request.POST.get("full_name", "").strip()
    username = request.POST.get("username", "").strip()
    email = request.POST.get("email", "").strip()
    password = request.POST.get("password", "")
    confirm_password = request.POST.get("confirm_password", "")
    mobile_number = request.POST.get("mobile_number", "")
    role = request.POST.get("role", "operator")
    user_group = request.POST.get("user_group", "operations")
    tz = request.POST.get("timezone", "Asia/Kolkata")
    language = request.POST.get("language", "English")
    is_active_val = request.POST.get("is_active_user", "1")
    account_expires = request.POST.get("account_expires", "") or None
    remarks = request.POST.get("remarks", "")
    receive_email = "receive_email" in request.POST
    receive_sms = "receive_sms" in request.POST
    mobile_app_access = "mobile_app_access" in request.POST

    # Validation
    if not full_name or not username or not email:
        messages.error(request, "Full name, username, and email are required.")
        return None

    if not user_obj and not password:
        messages.error(request, "Password is required for new users.")
        return None

    if password and password != confirm_password:
        messages.error(request, "Passwords do not match.")
        return None

    if not user_obj and User.objects.filter(username=username).exists():
        messages.error(request, f"Username '{username}' already exists.")
        return None

    # Split full name
    name_parts = full_name.split(" ", 1)
    first_name = name_parts[0]
    last_name = name_parts[1] if len(name_parts) > 1 else ""

    # Create or update User
    if user_obj is None:
        user_obj = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
        )
    else:
        user_obj.email = email
        user_obj.first_name = first_name
        user_obj.last_name = last_name
        if password:
            user_obj.set_password(password)
        user_obj.save()

    # Create or update Profile
    if profile_obj is None:
        profile_obj, _ = UserProfile.objects.get_or_create(user=user_obj)

    profile_obj.mobile_number = mobile_number
    profile_obj.role = role
    profile_obj.user_group = user_group
    profile_obj.timezone = tz
    profile_obj.language = language
    profile_obj.is_active_user = (is_active_val == "1")
    profile_obj.account_expires = account_expires
    profile_obj.remarks = remarks
    profile_obj.receive_email = receive_email
    profile_obj.receive_sms = receive_sms
    profile_obj.save()

    # Save Assigned Tanks / Devices
    assigned_tanks = request.POST.getlist("assigned_tanks")
    profile_obj.assigned_tanks.set(Tank.objects.filter(tank_id__in=assigned_tanks))

    # Save Module Permissions
    for mod in MODULE_LIST:
        key = mod["key"]
        perm, _ = ModulePermission.objects.get_or_create(
            profile=profile_obj,
            module_name=key,
        )
        perm.can_view = f"perm_{key}_view" in request.POST
        perm.can_add = f"perm_{key}_add" in request.POST
        perm.can_edit = f"perm_{key}_edit" in request.POST
        perm.can_delete = f"perm_{key}_delete" in request.POST
        perm.can_export = f"perm_{key}_export" in request.POST
        perm.save()

    return user_obj


@login_required
@check_license_expiry
def users_add_view(request):
    """Page: Add new user."""
    if request.method == "POST":
        result = _save_user_from_post(request)
        if result:
            messages.success(request, f"User '{result.username}' created successfully!")
            return redirect("users_list")

    modules = _build_module_context()
    context = {
        "editing": False,
        "edit_data": {},
        "modules": modules,
        "current_time": timezone.now(),
        "user": request.user,
        "all_tanks": Tank.objects.all(),
        "assigned_tank_ids": [],
    }
    return render(request, "core/users_add.html", context)


@login_required
@check_license_expiry
def users_edit_view(request, pk):
    """Page: Edit existing user."""
    profile = get_object_or_404(UserProfile, pk=pk)
    target_user = profile.user

    if request.method == "POST":
        result = _save_user_from_post(request, user_obj=target_user, profile_obj=profile)
        if result:
            messages.success(request, f"User '{result.username}' updated successfully!")
            return redirect("users_list")

    modules = _build_module_context(profile)
    edit_data = {
        "full_name": target_user.get_full_name(),
        "username": target_user.username,
        "email": target_user.email,
        "mobile_number": profile.mobile_number,
        "role": profile.role,
        "user_group": profile.user_group,
        "timezone": profile.timezone,
        "language": profile.language,
        "is_active_user": profile.is_active_user,
        "account_expires": profile.account_expires,
        "remarks": profile.remarks,
        "receive_email": profile.receive_email,
        "receive_sms": profile.receive_sms,
        "mobile_app_access": profile.mobile_app_access,
    }

    context = {
        "editing": True,
        "edit_data": edit_data,
        "modules": modules,
        "current_time": timezone.now(),
        "user": request.user,
        "all_tanks": Tank.objects.all(),
        "assigned_tank_ids": list(profile.assigned_tanks.values_list("tank_id", flat=True)),
    }
    return render(request, "core/users_add.html", context)


@login_required
@check_license_expiry
def users_delete_view(request, pk):
    """Delete a user and their profile."""
    profile = get_object_or_404(UserProfile, pk=pk)
    username = profile.user.username
    profile.user.delete()  # Cascade deletes profile + permissions
    messages.success(request, f"User '{username}' deleted successfully.")
    return redirect("users_list")


@login_required
def settings_add_tank_view(request):
    """API endpoint to dynamically add a new tank if it successfully connects."""
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            data = request.POST

        def get_float(val, default):
            if val in (None, ""):
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default

        def get_int(val, default):
            if val in (None, ""):
                return default
            try:
                return int(val)
            except (ValueError, TypeError):
                return default

        tank_id = data.get("tank_id", "").strip()
        name = data.get("name", "").strip()
        capacity_kl = get_float(data.get("capacity"), 5.0)
        capacity_unit = (data.get("capacity_unit") or "KL").strip()
        widget_type = data.get("widget_type", "cylinder").strip()
        
        flow_rate_register = get_int(data.get("flow_rate_register"), 40001)
        
        total_volume_register_val = data.get("total_volume_register")
        if total_volume_register_val and str(total_volume_register_val).strip():
            total_volume_register = get_int(total_volume_register_val, 40003)
        else:
            total_volume_register = None
            
        flow_unit = (data.get("flow_unit") or "L/min").strip()
        
        total_unit_val = data.get("total_unit")
        if total_unit_val and str(total_unit_val).strip():
            total_unit = total_unit_val.strip()
        else:
            total_unit = None

        address_str = data.get("address", "").strip()
        if widget_type == "flow_meter":
            address_str = str(flow_rate_register)
        
        # Try to connect and read from this register to verify communication
        from .models import SerialConnectionConfig
        config = SerialConnectionConfig.objects.filter(profile_name="Tanks").first()
        if not config:
            config = SerialConnectionConfig.objects.create(profile_name="Tanks")
            
        com_port_param = data.get("com_port")
        if com_port_param:
            config.com_port = com_port_param.strip()
            config.save()

        capacity_kl = get_float(data.get("capacity"), 5.0)
        high_limit = get_float(data.get("high_limit"), 90.0)
        low_limit = get_float(data.get("low_limit"), 10.0)
        raw_zero = get_float(data.get("raw_zero"), 0.0)
        raw_span = get_float(data.get("raw_span"), 0.0)
        scanner_raw_zero = get_float(data.get("scanner_raw_zero"), 0.0)
        scanner_raw_span = get_float(data.get("scanner_raw_span"), 0.0)
        slave_id = get_int(data.get("slave_id"), config.slave_id if config else 1)
        
        function_code_param = data.get("function_code")
        if function_code_param not in (None, ""):
            try:
                function_code = int(function_code_param)
            except:
                try:
                    function_code = 4 if int(address_str) >= 30000 and int(address_str) < 40000 else 3
                except:
                    function_code = 3
        else:
            try:
                function_code = 4 if int(address_str) >= 30000 and int(address_str) < 40000 else 3
            except:
                function_code = 3

        data_type = (data.get("data_type") or "UInt16").strip()
        byte_order = (data.get("byte_order") or "ABCD").strip()
        scaling = get_float(data.get("scaling"), 1.0)
        offset = get_float(data.get("offset"), 0.0)
        unit = (data.get("unit") or "%").strip()
        error_accuracy = get_float(data.get("error_accuracy"), 0.0)
        
        alarm_enabled_val = data.get("alarm_enabled", True)
        if isinstance(alarm_enabled_val, str):
            alarm_enabled = alarm_enabled_val.lower() == 'true'
        else:
            alarm_enabled = bool(alarm_enabled_val)
        
        if not tank_id or not name or not address_str:
            return JsonResponse({"success": False, "error": "Tank ID, Name, and Register Address are required."})

        if not address_str.isdigit():
            return JsonResponse({"success": False, "error": "Register Address must be a valid number."})

        raw_addr = int(address_str)

        com_port = config.com_port

        # Verify port is available (don't read register — let the daemon handle live readings)
        comm_verified = False
        warning_msg = None
        if com_port == "SIMULATOR":
            comm_verified = True
            # Auto-initialize simulated registers
            from .models import SimulatedRegister
            if widget_type == "flow_meter":
                if flow_rate_register is not None:
                    SimulatedRegister.objects.get_or_create(register_address=flow_rate_register, defaults={"value": 15})
                if total_volume_register is not None:
                    SimulatedRegister.objects.get_or_create(register_address=total_volume_register, defaults={"value": 120})
            else:
                if raw_addr is not None:
                    SimulatedRegister.objects.get_or_create(register_address=raw_addr, defaults={"value": 0})
        else:
            import serial
            try:
                import serial.tools.list_ports
                available_ports = [p.device for p in serial.tools.list_ports.comports()]
                if com_port in available_ports:
                    comm_verified = True
                else:
                    warning_msg = f"Widget added, but port {com_port} is not available. Device may be disconnected."
            except Exception as e:
                warning_msg = f"Widget added, but port check failed: {str(e)}. Device may be offline."

        # Create or update the tank/widget
        # Create or update the tank/widget
        from .models import FlowMeter
        if widget_type == "flow_meter" and FlowMeter.objects.filter(meter_id=tank_id).exists():
            fm = FlowMeter.objects.get(meter_id=tank_id)
            fm.name = name
            fm.flow_rate_register = flow_rate_register
            fm.total_volume_register = total_volume_register
            fm.flow_unit = flow_unit
            fm.total_unit = total_unit
            fm.slave_id = slave_id
            fm.high_limit = high_limit
            fm.low_limit = low_limit
            fm.error_accuracy = error_accuracy
            fm.alarm_enabled = alarm_enabled
            fm.save()

            if com_port_param:
                flow_config = SerialConnectionConfig.objects.filter(profile_name="Flow Meters").first()
                if not flow_config:
                    flow_config = SerialConnectionConfig.objects.create(profile_name="Flow Meters")
                flow_config.com_port = com_port_param.strip()
                flow_config.save()

            EventLog.objects.create(
                event_type="SYSTEM" if comm_verified else "WARNING",
                source="Settings",
                message=f"Updated Flow Meter '{name}' ({tank_id})."
            )
            return JsonResponse({"success": True, "message": f"Flow Meter '{name}' updated successfully!"})

        tank, created = Tank.objects.get_or_create(
            tank_id=tank_id,
            defaults={
                "name": name,
                "capacity_kl": capacity_kl,
                "capacity_unit": capacity_unit,
                "location": "Area 1",
                "is_active": True,
                "register_address": raw_addr,
                "widget_type": widget_type,
                "high_limit": high_limit,
                "low_limit": low_limit,
                "slave_id": slave_id,
                "function_code": function_code,
                "raw_zero": raw_zero,
                "raw_span": raw_span,
                "scanner_raw_zero": scanner_raw_zero,
                "scanner_raw_span": scanner_raw_span,
                "data_type": data_type,
                "byte_order": byte_order,
                "scaling": scaling,
                "offset": offset,
                "unit": unit,
                "flow_rate_register": flow_rate_register,
                "total_volume_register": total_volume_register,
                "flow_unit": flow_unit,
                "total_unit": total_unit,
                "error_accuracy": error_accuracy,
                "alarm_enabled": alarm_enabled,
            }
        )
        if not created:
            tank.name = name
            tank.register_address = raw_addr
            tank.widget_type = widget_type
            tank.capacity_kl = capacity_kl
            tank.capacity_unit = capacity_unit
            tank.high_limit = high_limit
            tank.low_limit = low_limit
            tank.slave_id = slave_id
            tank.function_code = function_code
            tank.raw_zero = raw_zero
            tank.raw_span = raw_span
            tank.scanner_raw_zero = scanner_raw_zero
            tank.scanner_raw_span = scanner_raw_span
            tank.data_type = data_type
            tank.byte_order = byte_order
            tank.scaling = scaling
            tank.offset = offset
            tank.unit = unit
            tank.flow_rate_register = flow_rate_register
            tank.total_volume_register = total_volume_register
            tank.flow_unit = flow_unit
            tank.total_unit = total_unit
            tank.error_accuracy = error_accuracy
            tank.alarm_enabled = alarm_enabled
            tank.save()

        # Immediately create/re-calculate TankReading if simulated value exists or previous raw_value exists
        from .models import SimulatedRegister
        sim_reg = SimulatedRegister.objects.filter(register_address=raw_addr).first()
        if sim_reg is not None:
            from .modbus_daemon import compute_tank_level_and_raw
            res = compute_tank_level_and_raw(tank, sim_reg.value)
            TankReading.objects.create(
                tank=tank,
                level_percent=res["level_percent"],
                raw_value=res["raw_value"]
            )
        else:
            prev_reading = tank.readings.first()
            if prev_reading and prev_reading.raw_value is not None:
                from .modbus_daemon import compute_tank_level_and_raw
                res = compute_tank_level_and_raw(tank, prev_reading.raw_value)
                TankReading.objects.create(
                    tank=tank,
                    level_percent=res["level_percent"],
                    raw_value=res["raw_value"]
                )

        # Log event
        EventLog.objects.create(
            event_type="SYSTEM" if comm_verified else "WARNING",
            source="Settings",
            message=f"Added/Updated Widget '{name}' ({tank_id}) with register {raw_addr}."
        )

        if warning_msg:
            resp_msg = warning_msg
        else:
            resp_msg = f"Widget '{name}' saved successfully!"
        return JsonResponse({"success": True, "message": resp_msg})

    return JsonResponse({"success": False, "error": "Only POST requests are allowed."})


@login_required
def settings_delete_tank_view(request, tank_id):
    """API endpoint to dynamically delete a tank."""
    tank = get_object_or_404(Tank, tank_id=tank_id)
    name = tank.name
    tank.delete() # Deletes tank + readings + alarms cascades
    
    # Log event
    EventLog.objects.create(
        event_type="SYSTEM",
        source="Settings",
        message=f"Deleted Tank '{name}' ({tank_id}) and all associated historical records."
    )

    return JsonResponse({"success": True, "message": f"Tank '{name}' deleted successfully."})


@login_required
def settings_update_capacities_view(request):
    """API endpoint to bulk update tank capacities."""
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            return JsonResponse({"success": False, "error": "Invalid request data format."})
            
        capacities = data.get("capacities", {})
        for tank_id, cap_val in capacities.items():
            try:
                tank = Tank.objects.get(tank_id=tank_id)
                tank.capacity_kl = float(cap_val)
                tank.save()
                
                # Log event
                EventLog.objects.create(
                    event_type="SYSTEM",
                    source="Settings",
                    message=f"Updated capacity of Tank '{tank.name}' ({tank.tank_id}) to {tank.capacity_kl} KL."
                )
            except (Tank.DoesNotExist, ValueError, TypeError):
                continue
                
        return JsonResponse({"success": True, "message": "Tank capacities updated successfully."})
    return JsonResponse({"success": False, "error": "POST request required."})


@login_required
def set_alarm_limits_view(request):
    """API endpoint to dynamically configure alarm thresholds for a tank or flow meter."""
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            data = request.POST

        device_id = data.get("device_id", data.get("tank_id", "")).strip()
        device_type = data.get("device_type", "tank").strip()
        high_limit_val = data.get("high_limit")
        low_limit_val = data.get("low_limit")

        if not device_id:
            return JsonResponse({"success": False, "error": "Device ID is required."})

        try:
            if device_type == "flow_meter":
                device = get_object_or_404(FlowMeter, meter_id=device_id)
                unit_str = f" {device.flow_unit}"
            else:
                device = get_object_or_404(Tank, tank_id=device_id)
                unit_str = f"{device.unit}"

            device.high_limit = float(high_limit_val)
            device.low_limit = float(low_limit_val)
            device.save()
        except (ValueError, TypeError) as e:
            return JsonResponse({"success": False, "error": "Limits must be valid numbers."})

        # Log event
        EventLog.objects.create(
            event_type="SYSTEM",
            source="Alarms Config",
            message=f"Updated alarm limits for {device_type} '{device.name}' ({device_id}): High={device.high_limit}{unit_str}, Low={device.low_limit}{unit_str}."
        )

        return JsonResponse({"success": True, "message": f"Alarm settings for '{device.name}' updated successfully."})

    return JsonResponse({"success": False, "error": "Only POST requests are allowed."})


@login_required
def toggle_device_alarm_view(request):
    """API endpoint to toggle alarm_enabled for a tank or flow meter."""
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            data = request.POST

        device_id = data.get("device_id", "").strip()
        device_type = data.get("device_type", "tank").strip()
        enabled = data.get("enabled", True)

        if not device_id:
            return JsonResponse({"success": False, "error": "Device ID is required."})

        try:
            if device_type == "flow_meter":
                device = get_object_or_404(FlowMeter, meter_id=device_id)
            else:
                device = get_object_or_404(Tank, tank_id=device_id)

            device.alarm_enabled = bool(enabled)
            device.save()
            
            state = "enabled" if device.alarm_enabled else "disabled"
            msg = f"Alarms for {device_type} '{device.name}' ({device_id}) {state} successfully."
            
            EventLog.objects.create(
                event_type="SYSTEM",
                source="Alarms Config",
                message=msg
            )
            return JsonResponse({"success": True, "message": msg})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Only POST requests are allowed."})



@login_required
def settings_save_simulated_register_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except Exception:
            return JsonResponse({"success": False, "error": "Invalid request body."})
        
        try:
            register_address = int(data.get("register_address"))
            value = int(data.get("value"))
        except (ValueError, TypeError):
            return JsonResponse({"success": False, "error": "Register address and value must be valid integers."})
            
        from .models import SimulatedRegister, Tank, TankReading, FlowMeter, FlowMeterReading
        sim_reg, created = SimulatedRegister.objects.get_or_create(
            register_address=register_address,
            defaults={"value": value}
        )
        if not created:
            sim_reg.value = value
            sim_reg.save()

        # Update readings immediately for mapped Tanks & Flow Meters
        from .modbus_daemon import compute_tank_level_and_raw
        tanks = Tank.objects.filter(register_address=register_address, is_active=True)
        for tank in tanks:
            res = compute_tank_level_and_raw(tank, value)
            TankReading.objects.create(
                tank=tank,
                level_percent=res["level_percent"],
                raw_value=res["raw_value"]
            )
            
        flow_meters = FlowMeter.objects.filter(flow_rate_register=register_address, is_active=True)
        for fm in flow_meters:
            raw_flow = float(value)
            if fm.scanner_raw_span > fm.scanner_raw_zero:
                flow_rate = ((raw_flow - fm.scanner_raw_zero) / (fm.scanner_raw_span - fm.scanner_raw_zero)) * fm.calibrated_span
                flow_rate = flow_rate + fm.error_accuracy
            else:
                flow_rate = raw_flow + fm.error_accuracy
            FlowMeterReading.objects.create(
                flow_meter=fm,
                flow_rate=max(0.0, flow_rate),
                total_volume=0.0
            )

        # Create a system log for settings change
        EventLog.objects.create(
            event_type="SYSTEM",
            source="Simulator",
            message=f"Simulated register {register_address} value set to {value}."
        )
        return JsonResponse({"success": True, "message": f"Register {register_address} updated successfully."})
        
    return JsonResponse({"success": False, "error": "Method not allowed."})


@login_required
def settings_auto_port_scan_view(request):
    import serial.tools.list_ports
    try:
        # Return ONLY physical COM ports (no SIMULATOR) for scan results
        ports = [p.device for p in serial.tools.list_ports.comports()]
        ports = sorted(list(set(ports)))
        return JsonResponse({"success": True, "ports": ports})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
def settings_auto_slave_scan_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            data = request.POST
        
        com_port = data.get("com_port", "COM3")
        start_id = int(data.get("start_id", 1))
        end_id = int(data.get("end_id", 15))
        
        start_id = max(1, min(247, start_id))
        end_id = max(1, min(247, max(start_id, end_id)))
        if end_id - start_id > 30:
            end_id = start_id + 30
            
        if com_port == "SIMULATOR":
            return JsonResponse({"success": True, "slaves": [1, 2, 5]})
            
        import serial
        import minimalmodbus
        
        active_slaves = []
        try:
            import serial.tools.list_ports
            available_ports = [p.device for p in serial.tools.list_ports.comports()]
            if com_port not in available_ports:
                return JsonResponse({"success": False, "error": f"Port {com_port} is not available."})
        except:
            return JsonResponse({"success": False, "error": "Serial ports listing failed."})
            
        from .models import SerialConnectionConfig
        # Find the config that matches this COM port, or fall back to any config
        config = SerialConnectionConfig.objects.filter(com_port=com_port).first()
        if not config:
            config = SerialConnectionConfig.objects.first()
        baudrate = config.baud_rate if config else 9600
        bytesize = config.data_bits if config else 8
        parity_str = config.parity if config else "None"
        stopbits = config.stop_bits if config else 1
        
        parity_map = {
            "None": serial.PARITY_NONE,
            "Even": serial.PARITY_EVEN,
            "Odd": serial.PARITY_ODD,
        }
        parity = parity_map.get(parity_str, serial.PARITY_NONE)
        
        # Find all unique baud rates in database, or default
        bauds = list(SerialConnectionConfig.objects.values_list('baud_rate', flat=True).distinct())
        if not bauds:
            bauds = [9600, 19200]
            
        from .modbus_daemon import get_com_lock
        with get_com_lock(com_port):
            active_slaves = []
            for test_baud in bauds:
                instrument = None
                try:
                    instrument = minimalmodbus.Instrument(com_port, start_id, close_port_after_each_call=False)
                    instrument.serial.baudrate = test_baud
                    instrument.serial.bytesize = bytesize
                    instrument.serial.parity = parity
                    instrument.serial.stopbits = stopbits
                    instrument.serial.timeout = 0.3
                    instrument.clear_buffers_before_each_transaction = True
                    
                    for slave_id in range(start_id, end_id + 1):
                        if slave_id in active_slaves:
                            continue
                            
                        instrument.address = slave_id
                        responded = False
                        
                        for fc in [3, 4]:
                            for addr in [0, 1]:
                                # Try 32-bit first
                                try:
                                    instrument.read_registers(addr, 2, functioncode=fc)
                                    responded = True
                                    break
                                except Exception as fc_err:
                                    err_str = str(fc_err).lower()
                                    if "slave reported" in err_str or "exception code" in err_str or "illegal data address" in err_str:
                                        responded = True
                                        break
                                    elif "SlaveReportedException" in fc_err.__class__.__name__:
                                        responded = True
                                        break
                                        
                                # Try 16-bit
                                try:
                                    instrument.read_register(addr, functioncode=fc)
                                    responded = True
                                    break
                                except Exception as fc_err:
                                    err_str = str(fc_err).lower()
                                    if "slave reported" in err_str or "exception code" in err_str or "illegal data address" in err_str:
                                        responded = True
                                        break
                                    elif "SlaveReportedException" in fc_err.__class__.__name__:
                                        responded = True
                                        break
                            if responded:
                                break
                                
                        if responded:
                            active_slaves.append(slave_id)
                except Exception as scan_err:
                    pass
                finally:
                    if instrument and instrument.serial and instrument.serial.is_open:
                        try:
                            instrument.serial.close()
                        except:
                            pass
                            
        return JsonResponse({"success": True, "slaves": active_slaves})
        
    return JsonResponse({"success": False, "error": "POST request expected."})


@login_required
def settings_register_scan_view(request):
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body)
        except:
            data = request.POST
            
        com_port = data.get("com_port", "COM3")
        slave_id = int(data.get("slave_id", 1))
        auto_scan = data.get("auto_scan", False)
        data_type = data.get("data_type", "UInt16")
        byte_order = data.get("byte_order", "ABCD")
        
        if auto_scan in [True, "true", "True", 1, "1"]:
            start_reg = 40001
            end_reg = 40050
            func_code = 3
        else:
            start_reg = int(data.get("start_register", 40001))
            end_reg = int(data.get("end_register", 40015))
            func_code = int(data.get("function_code", 3))
            
            start_reg = max(0, start_reg)
            end_reg = max(0, max(start_reg, end_reg))
            if end_reg - start_reg > 30:
                end_reg = start_reg + 30
            
        if com_port == "SIMULATOR":
            from .models import SimulatedRegister
            regs = SimulatedRegister.objects.filter(register_address__range=(start_reg, end_reg))
            results = []
            for r in regs:
                results.append({"register": r.register_address, "value": r.value, "status": "Success"})
            return JsonResponse({"success": True, "results": results, "baud_rate": "SIMULATED"})
            
        import serial
        import minimalmodbus
        
        try:
            import serial.tools.list_ports
            available_ports = [p.device for p in serial.tools.list_ports.comports()]
            if com_port not in available_ports:
                return JsonResponse({"success": False, "error": f"Port {com_port} is not available."})
        except:
            return JsonResponse({"success": False, "error": "Serial ports listing failed."})
            
        from .models import SerialConnectionConfig
        config = SerialConnectionConfig.objects.first()
        config_baud = config.baud_rate if config else 9600
        bytesize = config.data_bits if config else 8
        parity_str = config.parity if config else "None"
        stopbits = config.stop_bits if config else 1
        
        parity_map = {
            "None": serial.PARITY_NONE,
            "Even": serial.PARITY_EVEN,
            "Odd": serial.PARITY_ODD,
        }
        parity = parity_map.get(parity_str, serial.PARITY_NONE)
        
        bauds_to_test = [config_baud, 9600, 19200, 115200] if auto_scan else [config_baud]
        bauds_to_test = list(dict.fromkeys(bauds_to_test))
        
        results = []
        found_baud = None
        from .modbus_daemon import get_com_lock, decode_registers
        with get_com_lock(com_port):
            instrument = None
            try:
                if auto_scan:
                    # 1. Detect Baud Rate using a quick read on register 0/1
                    for test_baud in bauds_to_test:
                        try:
                            instrument = minimalmodbus.Instrument(com_port, slave_id, close_port_after_each_call=True)
                            instrument.serial.baudrate = test_baud
                            instrument.serial.bytesize = bytesize
                            instrument.serial.parity = parity
                            instrument.serial.stopbits = stopbits
                            instrument.serial.timeout = 0.3
                            instrument.clear_buffers_before_each_transaction = True
                            
                            for addr in [0, 1]:
                                try:
                                    instrument.read_register(addr, functioncode=func_code)
                                    found_baud = test_baud
                                    break
                                except minimalmodbus.SlaveReportedException:
                                    found_baud = test_baud
                                    break
                                except Exception as e:
                                    err_str = str(e).lower()
                                    if "slave reported" in err_str or "illegal data address" in err_str:
                                        found_baud = test_baud
                                        break
                            if found_baud:
                                break
                        except Exception:
                            pass
                            
                    if not found_baud:
                        return JsonResponse({"success": False, "error": "Could not detect baud rate. No device responded on the selected slave ID."})
                        
                    # 2. Scan a broader range
                    instrument = minimalmodbus.Instrument(com_port, slave_id, close_port_after_each_call=False)
                    instrument.serial.baudrate = found_baud
                    instrument.serial.bytesize = bytesize
                    instrument.serial.parity = parity
                    instrument.serial.stopbits = stopbits
                    instrument.serial.timeout = 0.5
                    instrument.clear_buffers_before_each_transaction = True
                    
                    consecutive_zeros = 0
                    max_registers_to_scan = 100
                    
                    scan_start = int(data.get("start_register", 40001))
                    scan_end = scan_start + max_registers_to_scan
                    
                    num_regs = 1
                    if data_type in ("Int32", "UInt32", "Float32"):
                        num_regs = 2
                    elif data_type == "Float64":
                        num_regs = 4
                        
                    reg = scan_start
                    while reg < scan_end:
                        protocol_addr = reg - 40001 if reg >= 40000 else (reg - 30001 if reg >= 30000 else reg)
                        try:
                            regs = instrument.read_registers(protocol_addr, num_regs, functioncode=func_code)
                            val = decode_registers(regs, data_type, byte_order)
                            
                            if isinstance(val, float):
                                if val.is_integer():
                                    val = int(val)
                                else:
                                    val = round(val, 3)
                                    
                            results.append({"register": reg, "value": val, "status": "Success"})
                            
                            if val == 0:
                                consecutive_zeros += 1
                            else:
                                consecutive_zeros = 0
                                
                            if consecutive_zeros >= 5:
                                results = results[:-5]
                                break
                        except Exception as e:
                            err_str = str(e).lower()
                            if "illegal data address" in err_str or "slave reported" in err_str:
                                break
                            pass
                        reg += num_regs
                else:
                    # Manual Scan logic
                    found_baud = config_baud
                    instrument = minimalmodbus.Instrument(com_port, slave_id, close_port_after_each_call=False)
                    instrument.serial.baudrate = found_baud
                    instrument.serial.bytesize = bytesize
                    instrument.serial.parity = parity
                    instrument.serial.stopbits = stopbits
                    instrument.serial.timeout = 0.5
                    instrument.clear_buffers_before_each_transaction = True
                    
                    num_regs = 1
                    if data_type in ("Int32", "UInt32", "Float32"):
                        num_regs = 2
                    elif data_type == "Float64":
                        num_regs = 4
                        
                    reg = start_reg
                    while reg <= end_reg:
                        protocol_addr = reg - 40001 if reg >= 40000 else (reg - 30001 if reg >= 30000 else reg)
                        try:
                            regs = instrument.read_registers(protocol_addr, num_regs, functioncode=func_code)
                            val = decode_registers(regs, data_type, byte_order)
                            
                            if isinstance(val, float):
                                if val.is_integer():
                                    val = int(val)
                                else:
                                    val = round(val, 3)
                                    
                            results.append({"register": reg, "value": val, "status": "Success"})
                        except Exception as e:
                            results.append({"register": reg, "value": None, "status": f"Failed: {str(e)}"})
                        reg += num_regs
            except Exception as scan_err:
                return JsonResponse({"success": False, "error": f"Scanner error: {str(scan_err)}"})
            finally:
                if instrument and instrument.serial and instrument.serial.is_open:
                    try:
                        instrument.serial.close()
                    except:
                        pass
                    
        return JsonResponse({"success": True, "results": results, "baud_rate": found_baud})
        
    return JsonResponse({"success": False, "error": "POST request expected."})


@login_required
def api_diagnostics_view(request):
    from .models import ModbusDiagnosticStats, ModbusFrameLog
    stats, _ = ModbusDiagnosticStats.objects.get_or_create(id=1)
    
    frames = list(ModbusFrameLog.objects.all()[:30])
    frame_list = []
    for f in frames:
        frame_list.append({
            "timestamp": f.timestamp.strftime("%H:%M:%S.%f")[:-3],
            "direction": f.direction,
            "frame_hex": f.frame_hex
        })
        
    return JsonResponse({
        "success": True,
        "stats": {
            "tx_packets": stats.tx_packets,
            "rx_packets": stats.rx_packets,
            "crc_errors": stats.crc_errors,
            "timeout_count": stats.timeout_count,
            "response_time_ms": round(stats.response_time_ms, 1),
            "comm_quality": round(stats.comm_quality, 1),
            "last_comm": stats.last_comm.strftime("%Y-%m-%d %I:%M:%S %p") if stats.last_comm else "Never"
        },
        "frames": frame_list
    })

@login_required
def api_active_alarms(request):
    """Returns the first unacknowledged alarm to show as a popup."""
    # Find oldest unacknowledged alarm
    alarm = Alarm.objects.filter(acknowledged=False).order_by('timestamp').first()
    if alarm:
        device_name = alarm.tank.name if alarm.tank else (alarm.flow_meter.name if alarm.flow_meter else "Unknown Device")
        device_id = alarm.tank.tank_id if alarm.tank else (alarm.flow_meter.meter_id if alarm.flow_meter else "Unknown ID")
        return JsonResponse({
            "has_alarm": True,
            "alarm_id": alarm.id,
            "device_id": device_id,
            "device_name": device_name,
            "alarm_type": alarm.alarm_type, # "high" or "low"
            "level": alarm.level_percent,
            "timestamp": alarm.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        })
    return JsonResponse({"has_alarm": False})

@login_required
def api_acknowledge_alarm(request):
    """Toggles alarm acknowledgment status and updates snooze."""
    import json
    from datetime import timedelta
    from django.utils import timezone
    
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            alarm_id = data.get("alarm_id")
            alarm = Alarm.objects.get(id=alarm_id)
            alarm.acknowledged = not alarm.acknowledged
            if alarm.acknowledged:
                alarm.snooze_until = timezone.now() + timedelta(minutes=10)
            else:
                alarm.snooze_until = timezone.now() - timedelta(minutes=1)  # Clear snooze
            alarm.save()
            return JsonResponse({"status": "success", "acknowledged": alarm.acknowledged})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
    return JsonResponse({"status": "error", "message": "Invalid request method"})


@login_required
def api_alarm_detail(request, alarm_id):
    """Returns detailed JSON information for a specific alarm."""
    from django.utils import timezone
    alarm = get_object_or_404(Alarm, id=alarm_id)
    device_name = alarm.tank.name if alarm.tank else (alarm.flow_meter.name if alarm.flow_meter else "Unknown Device")
    device_id = alarm.tank.tank_id if alarm.tank else (alarm.flow_meter.meter_id if alarm.flow_meter else "Unknown ID")
    
    if alarm.tank:
        high_limit = alarm.tank.high_limit
        low_limit = alarm.tank.low_limit
        unit = "%"
        formatted_level = f"{alarm.level_percent:.1f}%"
        dev_type = "Tank Level"
    elif alarm.flow_meter:
        high_limit = alarm.flow_meter.high_limit
        low_limit = alarm.flow_meter.low_limit
        unit = f" {alarm.flow_meter.flow_unit}"
        formatted_level = f"{alarm.level_percent:.1f}{unit}"
        dev_type = "Flow Meter"
    else:
        high_limit = low_limit = 0
        unit = ""
        formatted_level = f"{alarm.level_percent:.1f}"
        dev_type = "Unknown"

    message = f"Value exceeded High Limit ({high_limit}{unit})" if alarm.alarm_type == "high" else f"Value dropped below Low Limit ({low_limit}{unit})"

    return JsonResponse({
        "success": True,
        "alarm_id": alarm.id,
        "device_id": device_id,
        "device_name": device_name,
        "device_type": dev_type,
        "alarm_type": "High Alarm" if alarm.alarm_type == "high" else "Low Alarm",
        "severity": "HIGH" if alarm.alarm_type == "high" else "LOW",
        "level": formatted_level,
        "high_limit": f"{high_limit}{unit}",
        "low_limit": f"{low_limit}{unit}",
        "message": message,
        "timestamp": timezone.localtime(alarm.timestamp).strftime("%d %b %Y %I:%M:%S %p"),
        "acknowledged": alarm.acknowledged,
        "status": "Acknowledged" if alarm.acknowledged else "Active"
    })
